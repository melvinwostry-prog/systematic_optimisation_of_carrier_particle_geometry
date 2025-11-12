


import os
import openpyxl
import numpy as np
import random
import math
from openpyxl.styles import Alignment
from formlayout import fedit
import time
from datetime import datetime


def Empty_Inputs():
    global Array_dict
    Array_dict = {}
    global P_Array_dict
    P_Array_dict = {}
    global R_Array_dict
    R_Array_dict = {}
    global FollowUp
    FollowUp = True
    global Simulation_Letter
    Simulation_Letter = 'B'

Empty_Inputs()
    

def Define_Path():
    os.chdir('D:\Rocky_Simulations\Empty_Simulation_Folder\Process')
    wb = openpyxl.load_workbook('Recent.xlsx')
    Carriersheet = wb['Path']
    Recent_Path = Carriersheet['A1'].value
    Recent_Iterations = Carriersheet['A2'].value
    Recent_Carrier_Replicas = Carriersheet['A3'].value
    Recent_Replicas = Carriersheet['A4'].value
    Option_Choice = [
        ('Path_Number',Recent_Path),
        ('Number of total Iterations',Recent_Iterations),
        ('Number of Replicas for Loading',Recent_Carrier_Replicas),
        ('Number of Replicas for Collision',Recent_Replicas)
        ]
    setup = fedit(Option_Choice, title='Simulationpath', 
                  comment='Which Simulationpath is in progress?'
                  '\nNo punctuation!')
    global Path_Number
    global max_iterations
    global Replicas
    global Loading_Replicas
    Path_Number_long = setup[0]
    max_iterations = setup[1]
    Loading_Replicas = setup[2]
    Replicas = setup[3]
    global n_max
    n_max = Replicas + 1 ##wieso +1?
    number_str = str(Path_Number_long)
    Path_Number = ".".join(number_str)
    Carriersheet['A1'] = Path_Number_long
    Carriersheet['A2'] = max_iterations
    Carriersheet['A3'] = Loading_Replicas
    Carriersheet['A4'] = Replicas
    wb.save('Recent.xlsx')

Define_Path()

def Update_Iterations():
    os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\DoE')
#    while True:
#        now = datetime.now()
#        if 30 <= now.second <= 40:
#            break
#        time.sleep(2)
    wb = openpyxl.load_workbook('DoE_Arrays.xlsx')
    #sheet = wb['Parameter']
    Resultsheet = wb['Parameter']
    last_col = Resultsheet.max_column
    filled_cells = 0
    for row in range(2, Resultsheet.max_row + 1):
        cell_value = Resultsheet.cell(row=row, column=last_col).value
        if cell_value is not None and str(cell_value).strip() != "":
            filled_cells += 1
    global Current_Iterations
    Current_Iterations = filled_cells

Update_Iterations()


def Report():
    os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
    wb = openpyxl.load_workbook('Carrier.xlsx')
    Carriersheet = wb['Carrier']
    for row in Carriersheet.iter_rows():
        for cell in row:
            if cell.font.bold:
                cell.font = openpyxl.styles.Font(bold=False)
    wb.save('Carrier.xlsx')

Report()

def DoE_Read():
    global Array_dict
    global Array_lenght
    #DoE
    os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\DoE')
    wb = openpyxl.load_workbook('DoE_Arrays.xlsx')
    sheet = wb['Parameter']
    Resultsheet = wb['Result']
    number_columns = sheet.max_column
    Array_lenght = sheet.max_row - 1
    for column_number in range(1, number_columns + 1):
        column_letter = openpyxl.utils.get_column_letter(column_number)
        Array_Name = sheet[column_letter + str(1)].value
        Array_value = np.array([(sheet.cell(row=row, column=column_number).value) for row in range(2, sheet.max_row + 1)])
        global Array_dict
        Array_dict[Array_Name] = Array_value
    number_columns = Resultsheet.max_column
    for column_number in range(1, number_columns + 1):
        column_letter = openpyxl.utils.get_column_letter(column_number)
        Array_Name = Resultsheet[column_letter + str(1)].value
        Array_value = np.array([(Resultsheet.cell(row=row, column=column_number).value) for row in range(2, Resultsheet.max_row + 1)])
        global R_Array_dict
        R_Array_dict[Array_Name] = Array_value
    return Array_lenght

DoE_Read()


def Parameter_Read():
    global P_Array_dict
    global P_Array_lenght
    #Parameter
    os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Parameter')
    Parameter_wb = openpyxl.load_workbook('Parameter.xlsx')
    Parameter_sheet = Parameter_wb['Parameter']
    Number_rows = Parameter_sheet.max_row
    #Array Lenght
    for Para in range(1,Parameter_sheet.max_column + 2):
        columnletter = openpyxl.utils.get_column_letter(Para)
        if Parameter_sheet[columnletter + '1'].value == None:
            global P_Array_lenght
            P_Array_lenght = Para - 1
            break
    #Array Einträge
    for rownumber in range(1, Number_rows + 1):
        P_Array_Name = Parameter_sheet['A' + str(rownumber)].value
        P_Array_value = np.array([Parameter_sheet.cell(row=rownumber, column=column).value for column in range(2, P_Array_lenght + 2)])
        P_Array_dict[P_Array_Name] = P_Array_value
            
Parameter_Read()


def Span():
    os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\DoE')
    wb = openpyxl.load_workbook('DoE_Arrays.xlsx')
    sheet = wb['Parameter']
    global Simulation_Array
    global Finished_Array
    global Simulation_Letter
    Simulation_Letter = 'G'
    Simulation_Array = []
    for Excel_row in range(2,sheet.max_row+1):
        if sheet['A' + str(Excel_row)].value != None:
            Simulation_Array.append(Excel_row)#((sheet['A' + str(Excel_row)].value)+1)
    os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
    wb = openpyxl.load_workbook('Carrier.xlsx')
    Carriersheet = wb['Carrier']
    Finished_Array = []
    for Excel_row in range(2,Carriersheet.max_row+1):
        if Carriersheet[Simulation_Letter + str(Excel_row)].value != None:
            Finished_Array.append([Excel_row])
    global Range
    Range = np.setdiff1d(Simulation_Array, Finished_Array)


def Geometry_Check():
    global CarrierGeometry
    global Geometry_Created
    os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Geometrien\Carrier')
    if os.path.isfile(CarrierGeometry + '.stl'):
        Geometry_Created = True
    else:
        Geometry_Created = False

def Break_Check():
   os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')     
   wb = openpyxl.load_workbook('Recent.xlsx')
   sheet = wb['Processes']
   Quit = sheet['B10'].value
   global Break_True
   Break_True = False
   if Quit != 0:
       Break_True = True

def Loading_Create():
    global Array_dict
    global P_Array_dict
    global Run_Opt
    global Range
    #Define Process
    Process = 'Loading'
    for ParaPosition in range(1,P_Array_lenght + 2):
        if P_Array_dict['Parameter'][ParaPosition - 1] == Process:
            Position = ParaPosition - 1
            break
    for Excel_row in Range:
        #input(str(Excel_row))
        Array_row = Excel_row - 2
        Run_row = Excel_row - 1
        #input(str(Run_row) + '_Run_row')
        os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
        wb = openpyxl.load_workbook('Carrier.xlsx')
        global FollowUp
        global Geometry_Created
        global CarrierGeometry
        Carriersheet = wb['Carrier']
        CarrierGeometry = str((Array_dict['Geometry'][Array_row]))
        Geometry_Check()
        if FollowUp == True and Geometry_Created == True:
            if Carriersheet['B' + str(Excel_row)].value == None or int(Carriersheet['B' + str(Excel_row)].value) < Loading_Replicas:
                if Carriersheet['B' + str(Excel_row)].value != None and Carriersheet['B' + str(Excel_row)].value < Loading_Replicas:
                    Start_Replica = int(Carriersheet['B' + str(Excel_row)].value + 1)
                else:
                    Start_Replica = 1
                CarrierGeometry = str((Array_dict['Geometry'][Array_row]))
                os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Loading\\Simulations')
                #Carrier Volume
                global CarrierParticleSize
                global Carrier_Volume
                #os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Geometrien\Carrier')
                #wb = openpyxl.load_workbook('Carrier_Parameter.xlsx')
                #sheet = wb['Parameter']    
                CarrierParticleSize = float(P_Array_dict['CarrierParticleSize'][Position]) * 10**(-6)   
                #for Carrier_row in range(2,sheet.max_row+1):
                #    Carrier_Name = str(sheet['A' + str(Carrier_row)].value)
                #    if Carrier_Name == CarrierGeometry:
                        #if sheet['B' + str(Carrier_row)].value != None:
                        #    Carrier_V_GH = sheet['B' + str(Carrier_row)].value * 1e-9
                        #    Carrier_D_GH = sheet['C' + str(Carrier_row)].value * 1e-3
                        #    Carrier_Volume = Carrier_V_GH * ((float(CarrierParticleSize) / Carrier_D_GH)**3)
                        #    break
                        #else:
                def Carrier_Volume_Search():
                    global Carrier_Volume
                    os.chdir('D:\Rocky_Simulations\Empty_Simulation_Folder\Geometrien\Carrier')
                    app.OpenProject('Carrier_Volume.rocky')
                    #Carrier Geometry laden
                    project = app.GetProject()
                    study = project.GetStudy()
                    ParticleCollection = study.GetParticleCollection()
                    CarrierParticle = ParticleCollection.New()
                    CarrierParticle.SetName('CarrierParticle')
                    os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Geometrien\Carrier')
                    CarrierParticle.SetShape('custom_polyhedron')
                    CarrierParticle.ImportFromSTL(CarrierGeometry + '.stl', 'as_concave')
                    #CarrierParticle.SetMaterial('Carrier')
                    Carrier_SizeDistribution = CarrierParticle.GetSizeDistributionList()
                    Carrier_SizeDistribution.Clear()
                    CarrierSize = Carrier_SizeDistribution.New()
                    CarrierSize.SetSize(CarrierParticleSize, 'm')
                    CarrierParticle.SetSuperquadricDegree(2.0)
                    CarrierParticle.SetEnableRotations(False)
                    #Volumetric inlet
                    study.CreateVolumetricInlet(CarrierParticle, 'CarrierInlet', 1e-11)
                    #Simulation laufen
                    study.StartSimulation(True, False)
                    #Partikel Volumen auslesen
                    user_processes = project.GetUserProcessCollection()
                    particles = study.GetParticles()
                    #Particle Mass
                    Particle_Mass_Process = user_processes.CreatePropertyProcess(particles, 'Properties')
                    Particle_Mass_Process.SetName('ParticleMassProcess')
                    Particle_Mass_Process.SetPropertyGridFunction('Particle Group')
                    Particle_Mass_Process.SetCutValue(0)
                    Carrier_Volume_M = float(Particle_Mass_Process.GetGridFunction('Particle Volume').GetArray(time_step=1)[0]) * 10e17
                    Carrier_Volume = Carrier_Volume_M * 1e-18
                    #Simulation schließen
                    project.CloseProject(False)
                    os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Loading\\Simulations')
                Carrier_Volume_Search()
                #break
                #Create Simulation by Number of Replicas
                for Loading_Replicate in range(Start_Replica,Loading_Replicas+1):
                    if FollowUp == True:
                        app.CreateProject()
                        project = app.GetProject()
                        study = project.GetStudy()
                        study.SetCustomerName(CarrierGeometry)
                        os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Loading\\Simulations')
                        project.SaveProject(str(Run_row) + '_' + CarrierGeometry + '_L' + str(Loading_Replicate) + '_Loading.rocky')
                        Contacts = study.GetContactData()
                        ParticleCollection = study.GetParticleCollection()
                        #Define Parameters
                        for ParaPosition in range(1,P_Array_lenght + 2):
                            if P_Array_dict['Parameter'][ParaPosition - 1] == Process:
                                Position = ParaPosition - 1
                                break
                        for Parameter in P_Array_dict:
                            Value = P_Array_dict[Parameter][Position]
                            if Value == 'DoE':
                                DoE_Array_Value = Array_dict[Parameter]
                                DoE_Value = DoE_Array_Value[Array_row]
                                P_Array_dict[Parameter][Position] = DoE_Value
                        #Define Simulation
                        #Physics
                        physics = study.GetPhysics()
                        physics.SetGravityYDirection(float(P_Array_dict['GravityYDirection'][Position]))
                        physics.SetGravityStartTime(float(P_Array_dict['GravityStartTime'][Position]))
                        physics.SetGravityStopTime(float(P_Array_dict['GravityStopTime'][Position]))
                        physics.SetNormalForceModel(str(P_Array_dict['NormalForceModel'][Position]))
                        physics.SetTangentialForceModel(str(P_Array_dict['TangentialForceModel'][Position]))
                        physics.SetAdhesionModel(str(P_Array_dict['AdhesionModel'][Position]))
                        physics.SetRollingResistanceModel(str(P_Array_dict['RollingResistanceModel'][Position]))
                        physics.SetNumericalSofteningFactor(float(P_Array_dict['NumericalSofteningFactor'][Position]))
                        #Modules activate
                        module_collection = study.GetModuleCollection()
                        GetContactsOverlapMonitor = module_collection.GetModule('Contacts Overlap Monitor')
                        GetContactsOverlapMonitor.DisableModule()
                        GetParticlesEnergySpectra = module_collection.GetModule('Particles Energy Spectra')
                        GetParticlesEnergySpectra.DisableModule()
                        GetSPHDensityMonitor = module_collection.GetModule('SPH Density Monitor')
                        GetSPHDensityMonitor.DisableModule()
                        GetInter_particle_Collision_Statistics = module_collection.GetModule('Inter-particle Collision Statistics')
                        GetInter_particle_Collision_Statistics.EnableModule()
                        GetInter_particle_Collision_Statistics.SetModuleProperty('Duration', True)
                        GetInter_particle_Collision_Statistics.SetModuleProperty('Frequency', True)
                        GetInter_particle_Collision_Statistics.SetModuleProperty('Normal Impact Velocity', True)
                        GetInter_particle_Collision_Statistics.SetModuleProperty('Tangential Impact Velocity', True)
                        GetInter_particle_Collision_Statistics.SetModuleProperty('Power', True)
                        GetInter_particle_Collision_Statistics.SetModuleProperty('Force', True)
                        #Materials
                        MaterialCollection = study.GetMaterialCollection()
                        APIMaterial = MaterialCollection.New()
                        APIMaterial.SetName('API')
                        APIMaterial.SetDensity(float(P_Array_dict['APIDensity'][Position]))
                        APIMaterial.SetYoungsModulus(float(P_Array_dict['APIYoungsModulus'][Position]))
                        APIMaterial.SetPoissonRatio(float(P_Array_dict['APIPoissonRatio'][Position]))
                        CarrierMaterial = MaterialCollection.New()
                        CarrierMaterial.SetName('Carrier')
                        CarrierDensity = float(P_Array_dict['CarrierDensity'][Position])
                        if str(P_Array_dict['MassEquivalent'][Position]) == 'True':
                            SphereRadius = (float(P_Array_dict['CarrierParticleSize'][Position]) / 2)# * 10**(-6)
                            SphereVolume = ((4/3) * math.pi * ((SphereRadius)**3))
                            VolumeFactor = SphereVolume / Carrier_Volume
                            CarrierDensity = CarrierDensity * VolumeFactor
                        CarrierMaterial.SetDensity(CarrierDensity)
                        CarrierMaterial.SetYoungsModulus(float(P_Array_dict['CarrierYoungsModulus'][Position]))
                        CarrierMaterial.SetPoissonRatio(float(P_Array_dict['CarrierPoissonRatio'][Position]))
                        #Material Interactions
                        material_API = study.GetElement('API')
                        material_Carrier = study.GetElement('Carrier')
                        interaction_collection = study.GetMaterialsInteractionCollection()
                        interaction = interaction_collection.GetMaterialsInteraction(material_API, material_Carrier)
                        SurfaceEnergy = float(P_Array_dict['APICarrierSurfaceEnergy'][Position]) * 0.001 ##Correction for Loading to minimise Overlap
                        interaction.SetSurfaceEnergy(SurfaceEnergy)
                        interaction.SetStaticFriction(float(P_Array_dict['APICarrierStaticFriction'][Position]))
                        interaction.SetDynamicFriction(float(P_Array_dict['APICarrierDynamicFriction'][Position]))
                        interaction.SetRestitutionCoefficient(float(P_Array_dict['APICarrierRestitutionCoefficient'][Position]))
                        interaction.SetContactStiffnessMultiplier(float(P_Array_dict['APICarrierContactStiffnessMultiplier'][Position]))
                        interaction = interaction_collection.GetMaterialsInteraction(material_API, material_API)
                        APIAPISurfaceEnergy = (float(P_Array_dict['APIAPISurfaceEnergy'][Position])) * 0.001
                        interaction.SetSurfaceEnergy(APIAPISurfaceEnergy)
                        interaction.SetStaticFriction(float(P_Array_dict['APIAPIStaticFriction'][Position]))
                        interaction.SetDynamicFriction(float(P_Array_dict['APIAPIDynamicFriction'][Position]))
                        interaction.SetRestitutionCoefficient(float(P_Array_dict['APIAPIRestitutionCoefficient'][Position]))
                        interaction.SetRestitutionCoefficient(0.1)
                        #Carrier Particle
                        ParticleCollection = study.GetParticleCollection()
                        CarrierParticle = ParticleCollection.New()
                        CarrierParticle.SetName('CarrierParticle')
                        os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Geometrien\Carrier')
                        CarrierParticle.SetShape('custom_polyhedron')
                        CarrierParticle.ImportFromSTL(CarrierGeometry + '.stl', 'as_concave')
                        CarrierParticle.SetMaterial('Carrier')
                        Carrier_SizeDistribution = CarrierParticle.GetSizeDistributionList()
                        Carrier_SizeDistribution.Clear()
                        CarrierSize = Carrier_SizeDistribution.New()
                        CarrierSize.SetSize(CarrierParticleSize, 'm')
                        CarrierParticle.SetSuperquadricDegree(2.0)
                        CarrierParticle.SetEnableRotations(False)
                        #API Particle
                        study.GetParticleCollection()
                        APIPartikel = ParticleCollection.New()
                        APIPartikel.SetName('API')
                        APIPartikel.SetMaterial('API')
                        APIPartikel.SetRollingResistance(float(P_Array_dict['APIParticleRollingResistance'][Position]))
                        API_SizeDistribution = APIPartikel.GetSizeDistributionList()
                        API_SizeDistribution.Clear()
                        APISize = API_SizeDistribution.New()
                        APISize.SetSize(float(P_Array_dict['APISize'][Position]), 'um')
                        #Carrier Input
                        input_collection = study.GetParticleInputCollection()
                        CarrierInput = input_collection.AddCustomInput()
                        CarrierInput.SetName('Carrier Input')
                        os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Input Files')
                        if float(CarrierParticleSize) != 50:
                            wb = openpyxl.load_workbook('Carrier Custom Input File Rocky.xlsx')
                            sheet = wb['Klein Fix']
                            sheet['D2'] = CarrierParticleSize
                            wb.save('Carrier Custom Input File Rocky.xlsx')
                        CarrierInput.SetFilePath('Carrier Custom Input File Rocky.xlsx')
                        CarrierInput.SetParticle(CarrierParticle)
                        #Custom API Input
                        API_Size = float(P_Array_dict['APISize'][Position])
                        mean = API_Size
                        Range = 2
                        Faktor = 2
                        API_Upper_Limit = (2/3 * API_Size) + API_Size
                        API_Lower_Limit = API_Size - (2/3 * API_Size)
                        SampleNo = str(Excel_row-1)
                        std_dv = Range / Faktor
                        os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Input Files')
                        percentage = float(P_Array_dict['APILoad'][Position])
                        Maximum = 0
                        Carrier_Radius = (CarrierParticleSize / 2)
                        if str(P_Array_dict['MassEquivalent'][Position]) == 'True':
                            Carrier_Volume = SphereVolume
                        mCarrier = Carrier_Volume * float(P_Array_dict['CarrierDensity'][Position])
                        mAPI_Soll = percentage * 0.01 * mCarrier
                        mAPI_Ist = 0
                        Counter = 1
                        wb = openpyxl.load_workbook('API Default Custom Input File Rocky.xlsx')
                        sheet = wb['Parameter']
                        Maximum_Array = []
                        while mAPI_Ist < mAPI_Soll:
                            API_Size = 0
                            while API_Size < API_Lower_Limit or API_Upper_Limit < API_Size:
                                sample = np.random.normal(mean, std_dv)
                                API_Size = round(sample,2)
                            API_Radius = (API_Size / 2)
                            API_Volume = ((4/3) * math.pi * ((API_Radius)**3))
                            mAPI = API_Volume * 10**(-18) * float(P_Array_dict['APIDensity'][Position])
                            Maximum_Array.append(mAPI)
                            MaxRange = int(float(P_Array_dict['DomainXMax'][Position]) * 10 ** 9 * 0.95)
                            MinRange = int(float(Carrier_Radius * 10 ** 9))
                            x = 0
                            y = 0
                            z = 0
                            px = x
                            py = y
                            pz = z
                            while x ** 2  + y ** 2 + z ** 2 < MinRange ** 2 or x ** 2 > MaxRange ** 2 or y ** 2 > MaxRange ** 2 or z ** 2 > MaxRange ** 2:
                                x = random.randrange(-MaxRange, MaxRange)
                                y = random.randrange(-MaxRange, MaxRange)
                                z = random.randrange(-MaxRange, MaxRange)
                                if MinRange ** 2 <= x ** 2  + y ** 2 + z ** 2 and x ** 2 < MaxRange ** 2 and y ** 2 < MaxRange ** 2 and z ** 2 < MaxRange ** 2:
                                    px = x
                                    py = y
                                    pz = z
                            #Writing 
                            sheet['A' + str(Counter + 1)] = px * 10 ** -9
                            sheet['B' + str(Counter + 1)] = py * 10 ** -9
                            sheet['C' + str(Counter + 1)] = pz * 10 ** -9
                            sheet['D' + str(Counter + 1)] = API_Size * 10 ** -6
                            #Velocity
                            ux = 0
                            uy = 0
                            uz = 0
                            velocity = float(P_Array_dict['APIVelocity'][Position])
                            while (ux ** 2 + uy ** 2 + uz ** 2) ** 0.5 < 0.9 * velocity or (ux ** 2 + uy ** 2 + uz ** 2) ** 0.5 > 1.1 * velocity:
                                ux = random.uniform(-1, 1) * velocity
                                uy = random.uniform(-1, 1) * velocity
                                uz = random.uniform(-1, 1) * velocity
                            sheet['E' + str(Counter + 1)] = ux
                            sheet['F' + str(Counter + 1)] = uy
                            sheet['G' + str(Counter + 1)] = uz
                            mAPI_Ist = mAPI_Ist + mAPI
                            Counter = Counter + 1
                        wb.save('API_' + SampleNo + '_L' + str(Loading_Replicate) + '_Custom Input File Rocky.xlsx')
                        Maximum = sum(Maximum_Array)
                        #API Input
                        API_Input = input_collection.AddCustomInput()
                        API_Input.SetName('API_Input')
                        API_Input.SetFilePath('API_' + SampleNo + '_L' + str(Loading_Replicate) +  '_Custom Input File Rocky.xlsx')
                        API_Input.SetParticle(APIPartikel)
                        study.SetDescription(str(float(Maximum)))
                        #Contacts
                        Contacts.EnableCollectContactsData()
                        module_collection = study.GetModuleCollection()
                        GetContactsOverlapMonitor = module_collection.GetModule('Contacts Overlap Monitor')
                        GetContactsOverlapMonitor.DisableModule()
                        GetParticlesEnergySpectra = module_collection.GetModule('Particles Energy Spectra')
                        GetParticlesEnergySpectra.EnableModule()
                        GetParticlesEnergySpectra.SetModuleProperty('Impact Energy', True)
                        #Domain Settings
                        domain_settings = study.GetDomainSettings()
                        domain_settings.DisableUseBoundaryLimits()
                        domain_settings.SetCoordinateLimitsMinValues([float(P_Array_dict['DomainXMin'][Position]), float(P_Array_dict['DomainYMin'][Position]), float(P_Array_dict['DomainZMin'][Position])], 'm')
                        domain_settings.SetCoordinateLimitsMaxValues([float(P_Array_dict['DomainXMax'][Position]), float(P_Array_dict['DomainYMax'][Position]), float(P_Array_dict['DomainZMax'][Position])], 'm')
                        #Periodic Domain
                        domain_settings.SetDomainType('CARTESIAN')
                        domain_settings.SetCartesianPeriodicDirections('XYZ')
                        domain_settings.SetPeriodicLimitsMinCoordinates([float(P_Array_dict['PeriodicXMin'][Position]), float(P_Array_dict['PeriodicYMin'][Position]), float(P_Array_dict['PeriodicZMin'][Position])])
                        domain_settings.SetPeriodicLimitsMaxCoordinates([float(P_Array_dict['PeriodicXMax'][Position]), float(P_Array_dict['PeriodicYMax'][Position]), float(P_Array_dict['PeriodicZMax'][Position])])
                        domain_settings.SetPeriodicAtGeometryLimits(False)
                        #Solver
                        simulator_run = study.GetSimulatorRun()
                        simulator_run.SetSimulationDuration(float(P_Array_dict['SimulationDuration'][Position]))
                        simulator_run.SetSimulationOutputFrequency(float(P_Array_dict['SimulationOutputFrequency'][Position]))
                        simulator_run.SetSimulationTarget('CPU')
                        simulator_run.SetNumberOfProcessors(16)
                        os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Loading\\Simulations')
                        project.SaveProject()
                        #Autostart
                        os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
                        wb = openpyxl.load_workbook('Recent.xlsx')
                        Carriersheet = wb['Simulations']
                        Carriersheet['A1'] = str(Run_row) + '_' + CarrierGeometry + '_L' + str(Loading_Replicate) + 'Loading.rocky'
                        wb.save('Recent.xlsx')
                        #Simulation durchführen
                        os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Loading\\Simulations')
                        study.StartSimulation(True, False)
                        #Exit einbauen
                        time.sleep(3)
                        Resume = study.CanResumeSimulation()
                        if Resume == True:
                            FollowUp = False
                            project.CloseProject()
                            #break
                        if FollowUp == True:
                            os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
                            Carriersheet['A1'] = 'Empty'
                            wb.save('Recent.xlsx')
                            os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Loading\\Simulations')
                            project.SaveProject()
                            project.CloseProject()
                            os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
                            wb = openpyxl.load_workbook('Carrier.xlsx')
                            Carriersheet = wb['Carrier']
                            cell = Carriersheet['A' + str(Excel_row)]
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            Carriersheet['A' + str(Excel_row)] = (CarrierGeometry)
                            cell = Carriersheet['B' + str(Excel_row)]
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            Carriersheet['B' + str(Excel_row)] = Loading_Replicate
                            wb.save('Carrier.xlsx')
            else:
                continue
            

def Loading_Analyse():
    global Array_dict
    global P_Array_dict
    global Run_Opt
    global Range
    #Define Process
    Process = 'Loading'
    for ParaPosition in range(1,P_Array_lenght + 2):
        if P_Array_dict['Parameter'][ParaPosition - 1] == Process:
            Position = ParaPosition - 1
            break
    #Create Excel Files
    os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Loading\\Results')
    if not os.path.exists('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Loading\\Results\\Results_Loading.xlsx'):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Parameter'
        wb.save('Results_Loading.xlsx')
    wb = openpyxl.load_workbook('Results_Loading.xlsx')
    Parameter = wb['Parameter']
    Parameter['A1'] = 'Carrier'
    Parameter.merge_cells('A1:D1')
    Parameter['E1'] = 'API'
    Parameter.merge_cells('E1:H1')
    Parameter['A2'] = 'Geoemtry'
    Parameter['B2'] = 'Mass [ng]'
    Parameter['C2'] = 'Area [um2]'
    Parameter['D2'] = 'Volume [um3]'
    Parameter['E2'] = 'Diameter [µm]'
    Parameter['F2'] = 'Mass [ng]'
    Parameter['G2'] = 'Area [um2]'
    Parameter['H2'] = 'Volume [um3]'
    if not 'Total' in wb.sheetnames:
        wb.create_sheet('Total')
    sheet = wb['Total']
    sheet['A1'] = 'Geometry'
    sheet['B1'] = 'No'
    sheet['C1'] = 'Loading efficiency [%]'
    sheet['D1'] = 'Rel. St. Dv.'
    sheet['E1'] = '∅API Velocity [m/s]'
    if not 'Single' in wb.sheetnames:
        wb.create_sheet('Single')
    sheet = wb['Single']
    sheet['A1'] = 'Geometry'
    sheet['B1'] = 'Run'
    sheet['C1'] = 'Replicate'
    sheet['D1'] = 'Loading efficiency [%]'
    sheet['E1'] = '∅API Velocity [m/s]'
    wb.save('Results_Loading.xlsx')
    global Simulation_Letter
    #Simulation_Letter = 'C'
    Span()
    for Excel_row in Range:
        Array_row = Excel_row - 2
        Run_row = Excel_row - 1
        os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
        wb = openpyxl.load_workbook('Carrier.xlsx')
        Carriersheet = wb['Carrier']
        if Carriersheet['B' + str(Excel_row)].value == Loading_Replicas and Carriersheet['C' + str(Excel_row)].value == None:
            CarrierGeometry = str((Array_dict['Geometry'][Array_row]))
            #Define Parameters
            for Parameter in P_Array_dict:
                Value = P_Array_dict[Parameter][Position]
                if Value == 'DoE':
                    DoE_Array_Value = Array_dict[Parameter]
                    DoE_Value = DoE_Array_Value[Array_row]
                    P_Array_dict[Parameter][Position] = DoE_Value
            #Finished Simulation Check
            os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Loading\\Simulations')
            Total_Mass_Array = []
            Total_Velocity_Array = []
            for Loading_Replicate in range(1,Loading_Replicas+1):
                os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Loading\\Simulations')
                app.OpenProject(str(Run_row) + '_' + CarrierGeometry +  '_L' + str(Loading_Replicate) + '_Loading.rocky')        
                study = api.GetStudy()
                Finished = study.HasResults()
                Resume = study.CanResumeSimulation()
                project = api.GetProject()
                if Finished == False:
                    Decision = 'no'
                elif Finished == True:
                    Decision = 'yes'
                if Resume == True:
                    Decision = 'no'
                #Bei Entscheidung Ja
                if Decision == 'yes':
                    Sim_Dur = float(P_Array_dict['SimulationDuration'][Position])
                    Out_Fre = float(P_Array_dict['SimulationOutputFrequency'][Position])
                    Last_TS = Sim_Dur / Out_Fre
                    Last_TS = round(Last_TS)
                    project = app.GetProject()
                    study = project.GetStudy()
                    user_processes = project.GetUserProcessCollection()
                    particles = study.GetParticles()
                    #Maximum
                    Maximum = float(study.GetDescription()) * 10**12
                    #Velocity Carrier
                    Carrier_Particle_V = user_processes.CreatePropertyProcess(particles, 'Carrier_Particle_V')
                    Carrier_Particle_V.SetPropertyGridFunction('Particle Group')
                    Carrier_Particle_V.SetCutValue(0)
                    Carrier_V = float(Carrier_Particle_V.GetGridFunction('Absolute Translational Velocity').GetArray(time_step=Last_TS)[0])
                    API_Resting_V = Carrier_V * 1.1
                    #Velocity API
                    API_Particles = user_processes.CreatePropertyProcess(particles, 'API_Particles')
                    API_Particles.SetPropertyGridFunction('Particle Group')
                    API_Particles.SetCutValue(1)
                    API_0_V = user_processes.CreatePropertyProcess(API_Particles, '0_Velocity')
                    API_0_V.SetPropertyGridFunction('Absolute Translational Velocity')
                    API_0_V.SetType('Range')
                    API_0_V.SetMaxValue(0.01)
                    API_All_V = user_processes.CreatePropertyProcess(API_Particles, 'All_Velocity')
                    API_All_V.SetPropertyGridFunction('Absolute Translational Velocity')
                    API_All_V.SetType('Range')
                    API_All_V.SetMinValue(0)
                    API_All_V.SetMaxValue(100)
                    #Create Cube
                    Get_API = user_processes.GetProcess('0_Velocity')
                    Cube = user_processes.CreateCubeProcess(Get_API, 'Cube')
                    Cube.SetCenter(0.0, 0.0, 0.0, 'm')
                    Cube.SetSize(20.0, 10.0, 10.0, 'mm')
                    #Create Eulerian
                    Eulerian = user_processes.CreateEulerianStatistics(Cube, 'Eulerian')
                    Eulerian.SetDivisions((1, 1, 1,))
                    #Particle Mass
                    Particle_Mass_Process = user_processes.CreatePropertyProcess(particles, 'Properties')
                    Particle_Mass_Process.SetName('ParticleMassProcess')
                    Particle_Mass_Process.SetPropertyGridFunction('Particle Group')
                    Particle_Mass_Process.SetCutValue(0)
                    Carrier_Mass = float(Particle_Mass_Process.GetGridFunction('Particle Mass').GetArray(time_step=1)[0]) * 10e11
                    Carrier_SurfaceArea = float(Particle_Mass_Process.GetGridFunction('Particle Surface Area').GetArray(time_step=1)[0]) * 10e11
                    Carrier_Volume = float(Particle_Mass_Process.GetGridFunction('Particle Volume').GetArray(time_step=1)[0]) * 10e17
                    Particle_Mass_Process.SetCutValue(1)
                    Radius = float(P_Array_dict['APISize'][Position]) / 2
                    API_Volume = 4/3 * math.pi * Radius**3
                    API_Mass = (API_Volume * (float(P_Array_dict['APIDensity'][Position]))) / 10**6
                    API_SurfaceArea = 4 * math.pi * Radius ** 2
                    #Parameter
                    os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Loading\\Results')
                    wb = openpyxl.load_workbook('Results_Loading.xlsx')
                    Parameter = wb['Parameter']
                    spaltennummer = Excel_row
                    spaltennummer_API = 3
                    for Beta in range(3,Parameter.max_row + 2):
                        if Parameter['E' + str(Beta)] == float(P_Array_dict['APISize'][Position])  or Parameter['E' + str(spaltennummer)] != None:
                            spaltennummer_API = Beta
                            break
                    Parameter['A' + str(Excel_row+1)] = CarrierGeometry
                    Parameter['B' + str(Excel_row+1)] = Carrier_Mass
                    Parameter['C' + str(Excel_row+1)] = Carrier_SurfaceArea
                    Parameter['D' + str(Excel_row+1)] = Carrier_Volume
                    Parameter['E' + str(spaltennummer_API)] = float(P_Array_dict['APISize'][Position]) 
                    Parameter['F' + str(spaltennummer_API)] = API_Mass
                    Parameter['G' + str(spaltennummer_API)] = API_SurfaceArea
                    Parameter['H' + str(spaltennummer_API)] = API_Volume
                    wb.save('Results_Loading.xlsx')
                    #Carrier Position
                    Carrier = user_processes.CreatePropertyProcess(particles, 'Carrier')
                    Carrier.SetPropertyGridFunction('Particle Group')
                    Carrier.SetCutValue(0)
                    #Listen schreiben
                    Array_X = API_Particles.GetGridFunction('Particle X-Coordinate').GetArray(time_step=Last_TS)
                    Array_Y = API_Particles.GetGridFunction('Particle Y-Coordinate').GetArray(time_step=Last_TS)
                    Array_Z = API_Particles.GetGridFunction('Particle Z-Coordinate').GetArray(time_step=Last_TS)
                    Carrier_X_Orientation = Carrier_Particle_V.GetGridFunction('Particle X-Coordinate').GetArray(time_step=Last_TS)[0]
                    Carrier_Y_Orientation = Carrier_Particle_V.GetGridFunction('Particle Y-Coordinate').GetArray(time_step=Last_TS)[0]
                    Carrier_Z_Orientation = Carrier_Particle_V.GetGridFunction('Particle Z-Coordinate').GetArray(time_step=Last_TS)[0]
                    #API Maximale Resting Geschwindigkeit
                    Carrier_V = float(Carrier_Particle_V.GetGridFunction('Absolute Translational Velocity').GetArray(time_step=Last_TS)[0])
                    API_Resting_V = Carrier_V * 2
                    if API_Resting_V < 0.01:
                        API_Resting_V = 0.01
                    API_0_V.SetMaxValue(API_Resting_V)
                    n_API = float(Eulerian.GetGridFunction('Number of Particles').GetArray(time_step = Last_TS))  
                    #API Position, Alle API zum Last_TS durchscannen
                    No_API = 0
                    FaktorSize = float(P_Array_dict['APISize'][Position])
                    API_Mass_Array = []
                    API_SA_Array = []
                    for i in range(0,int(n_API)):
                        dx = Array_X[i] - Carrier_X_Orientation
                        dy = Array_Y[i] - Carrier_Y_Orientation
                        dz = Array_Z[i] - Carrier_Z_Orientation
                        dist = math.sqrt(dx**2 + dy**2 + dz**2)
                        if dist < FaktorSize:
                            No_API = No_API + 1
                            API_Single_Mass = float(API_0_V.GetGridFunction('Particle Mass').GetArray(time_step=Last_TS)[i]) * 10e11
                            API_Mass_Array.append(API_Single_Mass)
                            API_Diameter = float(API_0_V.GetGridFunction('Particle Size').GetArray(time_step=Last_TS)[i]) * 10e5
                            API_Radius = API_Diameter / 2
                            API_Area = math.pi * API_Radius ** 2
                            API_SA_Array.append(API_Area)    
                    ##Gesamtwerte ausrechnen  
                    API_Total_Mass = sum(API_Mass_Array)
                    API_Mass_Array.clear()
                    mRel = API_Total_Mass / Maximum
                    Velocity_Array = (API_Particles.GetGridFunction('Absolute Translational Velocity').GetArray(time_step=Last_TS))
                    Mean_Velocity = np.mean(Velocity_Array)
                    project.CloseProject(False)
                    #Ergebnisse Single schreiben
                    os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Loading\\Results')
                    wb = openpyxl.load_workbook('Results_Loading.xlsx')
                    sheet = wb['Single']
                    for row in range(2,sheet.max_row+2):
                        if sheet['A' + str(row)].value == None:
                            Single_row = row
                            break
                    sheet['A' + str(Single_row)] = CarrierGeometry
                    sheet['B' + str(Single_row)] = str(Run_row)
                    sheet['C' + str(Single_row)] = str(Loading_Replicate)
                    sheet['D' + str(Single_row)].number_format = '0.00%'
                    sheet['D' + str(Single_row)] = mRel
                    sheet['E' + str(Single_row)].number_format = '0.000'
                    sheet['E' + str(Single_row)] = Mean_Velocity
                    wb.save('Results_Loading.xlsx')
                    #Array für Ergebnisse Total vorbereiten
                    Total_Mass_Array.append(mRel)
                    Total_Velocity_Array.append(Mean_Velocity)
            #Ergebnisse Total schreiben
            os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Loading\\Results')
            wb = openpyxl.load_workbook('Results_Loading.xlsx')
            sheet = wb['Total']
            Total_row = Run_row + 1
            mRel = sum(Total_Mass_Array) / len(Total_Mass_Array)
            if Loading_Replicas != 1:
                varianz = sum((x - mRel) ** 2 for x in Total_Mass_Array) / (len(Total_Mass_Array) - 1)
                std_abw = varianz ** 0.5
                RelStDv = (std_abw / mRel)
            else:
                RelStDv = 'n.A.'
            Mean_Velocity = sum(Total_Velocity_Array) / len(Total_Velocity_Array)
            sheet['A' + str(Total_row)] = CarrierGeometry
            sheet['B' + str(Total_row)] = str(Run_row)
            sheet['C' + str(Total_row)].number_format = '0.00%'
            sheet['C' + str(Total_row)] = mRel
            sheet['D' + str(Total_row)].number_format = '0.00%'
            sheet['D' + str(Total_row)] = RelStDv
            sheet['E' + str(Total_row)].number_format = '0.000'
            sheet['E' + str(Total_row)] = Mean_Velocity
            wb.save('Results_Loading.xlsx')
            os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
            wb = openpyxl.load_workbook('Carrier.xlsx')
            Carriersheet = wb['Carrier']
            cell = Carriersheet['C' + str(Excel_row)]
            cell.font = cell.font.copy(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            Carriersheet['C' + str(Excel_row)] = 'X'
            wb.save('Carrier.xlsx')  
    #Spaltenbreite an Text anpassen
    #Parameter-sheet
    os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Loading\\Results')
    wb = openpyxl.load_workbook('Results_Loading.xlsx')
    sheet = wb['Parameter']
    for spalte in sheet.columns:
        max_laenge = 0
        spaltenbuchstabe = openpyxl.utils.get_column_letter(spalte[0].column)
        max_laenge = len(str(spalte[1].value))
        max_laenge2 = len(str(spalte[2].value))
        if max_laenge2 > max_laenge:
            max_laenge = max_laenge2
        adjusted_width = (max_laenge + 2) * 1
        if spaltenbuchstabe == 'A':
            adjusted_width *= 1.5
        sheet.column_dimensions[spaltenbuchstabe].width = adjusted_width
        wb.save('Results_Loading.xlsx')
    #Total-sheet
    sheet = wb['Total']
    for column in sheet.columns:
        spaltenbuchstabe = openpyxl.utils.get_column_letter(column[0].column)
        max_laenge = len(str(column[0].value))
        max_laenge2 = len(str(column[2].value))
        if max_laenge2 > max_laenge:
            max_laenge = max_laenge2
        adjusted_width = max_laenge + 2
        if spaltenbuchstabe == 'K':
            adjusted_width *= 1.5
        sheet.column_dimensions[spaltenbuchstabe].width = adjusted_width
    wb.save('Results_Loading.xlsx')

def ParticleParticle_Create():
    global Array_dict
    global P_Array_dict
    global Run_Opt
    global Range
    global FollowUp
    #Define Process
    Process = 'ParticleParticle'
    for ParaPosition in range(1,P_Array_lenght + 2):
        if P_Array_dict['Parameter'][ParaPosition - 1] == Process:
            Position = ParaPosition - 1
            break
    global Simulation_Letter
    Simulation_Letter = 'D'
    Span()
    for Excel_row in Range:
        Array_row = Excel_row - 2
        Run_row = Excel_row - 1
        os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
        wb = openpyxl.load_workbook('Carrier.xlsx')
        Carriersheet = wb['Carrier']
        Collision_Maximum = str(Loading_Replicas) + '.' + str(Replicas)
        if Carriersheet['C' + str(Excel_row)].value != None and Carriersheet['D' + str(Excel_row)].value != Collision_Maximum:
            #Finished Simulation Check
            CarrierGeometry = str((Array_dict['Geometry'][Array_row]))
            for Loading_Replicate in range(1,Loading_Replicas+1):
                os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Loading\\Simulations')
                if FollowUp == False:
                    break
                app.OpenProject(str(Run_row) + '_' + CarrierGeometry + '_L' + str(Loading_Replicate) + '_Loading.rocky')    
                study = api.GetStudy()
                Finished = study.HasResults()
                project = api.GetProject()
                if Finished == False:
                    Decision = 'no'
                elif Finished == True:
                    Decision = 'yes'
                if Decision == 'yes':
                    #Define Parameters
                    for ParaPosition in range(1,P_Array_lenght + 2):
                        if P_Array_dict['Parameter'][ParaPosition - 1] == Process:
                            Position = ParaPosition - 1
                            break
                    for Parameter in P_Array_dict:
                        Value = P_Array_dict[Parameter][Position]
                        if Value == 'DoE':
                            DoE_Array_Value = Array_dict[Parameter]
                            DoE_Value = DoE_Array_Value[Array_row]
                            P_Array_dict[Parameter][Position] = DoE_Value
                    CarrierSizeValue = float(P_Array_dict['CarrierParticleSize'][Position]) * 10**(-6)
                    Sim_Dur = float(P_Array_dict['SimulationDuration'][Position])
                    Out_Fre = float(P_Array_dict['SimulationOutputFrequency'][Position])
                    #Projekt
                    project = app.GetProject()
                    study = project.GetStudy()
                    user_processes = project.GetUserProcessCollection()
                    particles = study.GetParticles()
                    Contacts = study.GetContactData()
                    #Carrier
                    CarrierPosition = user_processes.CreatePropertyProcess(particles, 'Carrier')
                    CarrierPosition.SetPropertyGridFunction('Particle Group')
                    CarrierPosition.SetCutValue(0)
                    ##Direkte Kontakte
                    #Carrier
                    CarrierPosition = user_processes.CreatePropertyProcess(particles, 'CarrierPosition')
                    CarrierPosition.SetPropertyGridFunction('Particle Group')
                    CarrierPosition.SetCutValue(0)
                    #Velocity API
                    API_Particles = user_processes.CreatePropertyProcess(particles, 'API_Particles')
                    API_Particles.SetPropertyGridFunction('Particle Group')
                    API_Particles.SetCutValue(1)
                    API_0_V = user_processes.CreatePropertyProcess(API_Particles, '0_Velocity')
                    API_0_V.SetPropertyGridFunction('Absolute Translational Velocity')
                    API_0_V.SetType('Range')
                    API_0_V.SetMaxValue(float(P_Array_dict['APIVelocity'][Position]))
                    #Create Cube
                    Get_API = user_processes.GetProcess('0_Velocity')
                    Cube = user_processes.CreateCubeProcess(Get_API, 'Cube')
                    Cube.SetCenter(0.0, 0.0, 0.0, 'm')
                    Cube.SetSize(20.0, 10.0, 10.0, 'mm')
                    #Create Eulerian
                    Eulerian = user_processes.CreateEulerianStatistics(Cube, 'Eulerian')
                    Eulerian.SetDivisions((1, 1, 1,))
                    #Time Step
                    os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Loading\\Results')
                    wb = openpyxl.load_workbook('Results_Loading.xlsx')
                    sheet = wb['Total']
                    #Domain Bounds
                    Process = 'Loading'
                    for ParaPosition in range(1,P_Array_lenght + 2):
                        if P_Array_dict['Parameter'][ParaPosition - 1] == Process:
                            Position = ParaPosition - 1
                            break
                    DomainBounds = float(P_Array_dict['DomainXMax'][Position])
                    SimulationDuration = float(P_Array_dict['SimulationDuration'][Position])
                    OutputFrequency = float(P_Array_dict['SimulationOutputFrequency'][Position])
                    Process = 'ParticleParticle'
                    for ParaPosition in range(1,P_Array_lenght + 2):
                        if P_Array_dict['Parameter'][ParaPosition - 1] == Process:
                            Position = ParaPosition - 1
                            break
                    TS = int(SimulationDuration / OutputFrequency)
                    Carrier_Particle_V = user_processes.CreatePropertyProcess(particles, 'Carrier_Particle_V')
                    Carrier_Particle_V.SetPropertyGridFunction('Particle Group')
                    Carrier_Particle_V.SetCutValue(0)
                    Carrier_V = float(Carrier_Particle_V.GetGridFunction('Absolute Translational Velocity').GetArray(time_step=TS)[0])
                    API_Resting_V = Carrier_V * 2
                    API_0_V.SetMaxValue(API_Resting_V)  
                    for n in range(1,n_max):
                        #Write New input files
                        #Velocity Angle Division
                        Angle = float(P_Array_dict['CollisionAngle'][Position])
                        Velocity = float(P_Array_dict['CollisionVelocity'][Position])
                        Angle_in_radian = math.radians(Angle)
                        Velocity_X = Velocity * math.cos(Angle_in_radian)
                        Velocity_Y = Velocity * math.sin(Angle_in_radian)
                        for Partner in range(1,3):
                            #Carrier
                            Carrier_Angle = float(CarrierPosition.GetGridFunction('Orientation Angle').GetArray(time_step=TS))
                            Carrier_X_Orientation = float(CarrierPosition.GetGridFunction('Orientation Vector X').GetArray(time_step=TS))
                            Carrier_Y_Orientation = float(CarrierPosition.GetGridFunction('Orientation Vector Y').GetArray(time_step=TS))
                            Carrier_Z_Orientation = float(CarrierPosition.GetGridFunction('Orientation Vector Z').GetArray(time_step=TS))
                            os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Particle-Particle\\Custom Input Files')
                            wb = openpyxl.Workbook()
                            sheet = wb.active
                            sheet.title = 'Carrier'
                            Shift = 0
                            if Partner == 2:
                                Shift = 300e-06
                                Velocity_X = Velocity_X * (-1)
                            filename = (str(Run_row) + '_' + CarrierGeometry + '_Carrier Custom Input File Rocky P' + str(Partner) + '_L' + str(Loading_Replicate) + '_' + 'n' + str(n) + '.xlsx')
                            wb.save(filename)
                            wb = openpyxl.load_workbook(filename)
                            sheet = wb['Carrier']
                            sheet['A1'] = 'x'
                            sheet['B1'] = 'y'
                            sheet['C1'] = 'z'
                            sheet['A2'] = Shift
                            sheet['B2'] = 0
                            sheet['C2'] = 0
                            sheet['D1'] = 'size'
                            sheet['D2'] = CarrierSizeValue
                            sheet['F1'] = 'nx'
                            sheet['G1'] = 'ny'
                            sheet['H1'] = 'nz'
                            sheet['I1'] = 'ux'
                            sheet['J1'] = 'uy'
                            sheet['K1'] = 'uz'
                            sheet['I2'] = Velocity_X
                            sheet['J2'] = Velocity_Y
                            sheet['K2'] = 0
                            Angle = random.randint(0,90)
                            rad_Angle = Angle * (3.14159265359 / 180)
                            x_rotation = np.random.rand()
                            y_rotation = np.random.rand()
                            z_rotation = np.random.rand()
                            orientation_vector = np.array([x_rotation, y_rotation, z_rotation])
                            orientation_vector /= np.linalg.norm(orientation_vector)
                            sheet['E1'] = 'angle'
                            sheet['E2'] = rad_Angle
                            sheet['F2'] = orientation_vector[0]
                            sheet['G2'] = orientation_vector[1]
                            sheet['H2'] = orientation_vector[2]
                            wb.save(filename)
                            CarrierPosition_X = float(CarrierPosition.GetGridFunction('Particle X-Coordinate').GetArray(time_step=TS))
                            CarrierPosition_Y = float(CarrierPosition.GetGridFunction('Particle Y-Coordinate').GetArray(time_step=TS))
                            CarrierPosition_Z = float(CarrierPosition.GetGridFunction('Particle Z-Coordinate').GetArray(time_step=TS))
                            #API Position
                            API_X_array = API_0_V.GetGridFunction('Particle X-Coordinate').GetArray(time_step = TS)
                            API_Y_array = API_0_V.GetGridFunction('Particle Y-Coordinate').GetArray(time_step = TS)
                            API_Z_array = API_0_V.GetGridFunction('Particle Z-Coordinate').GetArray(time_step = TS)
                            #API
                            os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Particle-Particle\\Custom Input Files')
                            wb = openpyxl.Workbook()
                            sheet = wb.active
                            sheet.title = 'API'
                            filename = (str(Run_row) + '_' + CarrierGeometry + '_API Custom Input File Rocky P' + str(Partner) + '_L' + str(Loading_Replicate) + '_' + 'n' + str(n) + '.xlsx')
                            wb.save(filename)
                            wb = openpyxl.load_workbook(filename)
                            API_sheet = wb['API']
                            API_sheet['A1'] = 'x'
                            API_sheet['B1'] = 'y'
                            API_sheet['C1'] = 'z'
                            API_sheet['D1'] = 'size'
                            API_sheet['E1'] = 'ux'
                            API_sheet['F1'] = 'uy'
                            API_sheet['G1'] = 'uz'
                            numrows_1 = len(API_X_array)
                            Distance = 2 * DomainBounds
                            for a in range(0, numrows_1):
                                API_Diameter = float(API_0_V.GetGridFunction('Particle Size').GetArray(time_step=TS)[a])
                                #Shift Correction
                                #X
                                if (API_X_array[a] - CarrierPosition_X) > CarrierSizeValue:
                                    API_X_array[a] = API_X_array[a] - Distance
                                elif (API_X_array[a] - CarrierPosition_X) < - CarrierSizeValue:
                                    API_X_array[a] = API_X_array[a] + Distance
                                #Y
                                if (API_Y_array[a] - CarrierPosition_Y) > CarrierSizeValue:
                                    API_Y_array[a] = API_Y_array[a] - Distance
                                elif (API_Y_array[a] - CarrierPosition_Y) < - CarrierSizeValue:
                                    API_Y_array[a] = API_Y_array[a] + Distance
                                #Z
                                if (API_Z_array[a] - CarrierPosition_Z) > CarrierSizeValue:
                                    API_Z_array[a] = API_Z_array[a] - Distance
                                elif (API_Z_array[a] - CarrierPosition_Z) < - CarrierSizeValue:
                                    API_Z_array[a] = API_Z_array[a] + Distance
                                #Position
                                API_X_Shift = float(API_X_array[a]) - CarrierPosition_X
                                API_Y_Shift = float(API_Y_array[a]) - CarrierPosition_Y
                                API_Z_Shift = float(API_Z_array[a]) - CarrierPosition_Z
                                #erste Rotation
                                angle = Carrier_Angle
                                x_orient = Carrier_X_Orientation
                                y_orient = Carrier_Y_Orientation
                                z_orient = Carrier_Z_Orientation
                                P = np.array([API_X_Shift, API_Y_Shift, API_Z_Shift])
                                rotation_matrix = np.array([
                                [np.cos(angle) + x_orient**2 * (1 - np.cos(angle)),
                                 x_orient * y_orient * (1 - np.cos(angle)) - z_orient * np.sin(angle),
                                 x_orient * z_orient * (1 - np.cos(angle)) + y_orient * np.sin(angle)],
                                [y_orient * x_orient * (1 - np.cos(angle)) + z_orient * np.sin(angle),
                                 np.cos(angle) + y_orient**2 * (1 - np.cos(angle)),
                                 y_orient * z_orient * (1 - np.cos(angle)) - x_orient * np.sin(angle)],
                                [z_orient * x_orient * (1 - np.cos(angle)) - y_orient * np.sin(angle),
                                 z_orient * y_orient * (1 - np.cos(angle)) + x_orient * np.sin(angle),
                                 np.cos(angle) + z_orient**2 * (1 - np.cos(angle))]])
                                inverse_rotation_matrix = np.linalg.inv(rotation_matrix)
                                rotated_1_P = np.dot(inverse_rotation_matrix, P)
                                #zweite Rotation
                                angle = rad_Angle
                                x_orient = orientation_vector[0]
                                y_orient = orientation_vector[1]
                                z_orient = orientation_vector[2]
                                P = rotated_1_P
                                rotation_matrix = np.array([
                                [np.cos(angle) + x_orient**2 * (1 - np.cos(angle)),
                                 x_orient * y_orient * (1 - np.cos(angle)) - z_orient * np.sin(angle),
                                 x_orient * z_orient * (1 - np.cos(angle)) + y_orient * np.sin(angle)],
                                [y_orient * x_orient * (1 - np.cos(angle)) + z_orient * np.sin(angle),
                                 np.cos(angle) + y_orient**2 * (1 - np.cos(angle)),
                                 y_orient * z_orient * (1 - np.cos(angle)) - x_orient * np.sin(angle)],
                                [z_orient * x_orient * (1 - np.cos(angle)) - y_orient * np.sin(angle),
                                 z_orient * y_orient * (1 - np.cos(angle)) + x_orient * np.sin(angle),
                                 np.cos(angle) + z_orient**2 * (1 - np.cos(angle))]])
                                rotated_P = np.dot(rotation_matrix, P)
                                #Ausreißer entfernen
                                X_Position = rotated_P[0]
                                Y_Position = rotated_P[1]
                                Z_Position = rotated_P[2]
                                #X
                                if X_Position > CarrierSizeValue* 1.1 or X_Position < - CarrierSizeValue* 1.1:
                                    X_Position = (DomainBounds / 2) * 0.9
                                    Y_Position = (DomainBounds / 2) * 0.9
                                    Z_Position = (DomainBounds / 2) * 0.9
                                #Y
                                if Y_Position > CarrierSizeValue* 1.1 or Y_Position < - CarrierSizeValue* 1.1:
                                    X_Position = (DomainBounds / 2) * 0.9
                                    Y_Position = (DomainBounds / 2) * 0.9
                                    Z_Position = (DomainBounds / 2) * 0.9
                                #Z
                                if Z_Position > CarrierSizeValue* 1.1 or Z_Position < - CarrierSizeValue* 1.1:
                                    X_Position = (DomainBounds / 2) * 0.9
                                    Y_Position = (DomainBounds / 2) * 0.9
                                    Z_Position = (DomainBounds / 2) * 0.9
                                #Eintrag
                                API_sheet['A' + str(a + 2)] = X_Position + Shift
                                API_sheet['B' + str(a + 2)] = Y_Position
                                API_sheet['C' + str(a + 2)] = Z_Position
                                API_sheet['D' + str(a + 2)] = API_Diameter
                                API_sheet['E' + str(a + 2)] = Velocity_X
                                API_sheet['F' + str(a + 2)] = Velocity_Y
                                API_sheet['G' + str(a + 2)] = 0
                            wb.save(filename)
                    project.CloseProject(False)
                for n in range(1,n_max):
                    os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
                    wb = openpyxl.load_workbook('Carrier.xlsx')
                    Carriersheet = wb['Carrier']
                    Finished_n = Carriersheet['D' + str(Excel_row)].value
                    Collision_Maximum = str(Loading_Replicas) + '.' + str(Replicas)
                    if Finished_n == None or Finished_n != Collision_Maximum:
                        #Simulation erstellen
                        app.CreateProject()
                        project = app.GetProject()
                        study = project.GetStudy()
                        study.SetCustomerName(CarrierGeometry)
                        os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Particle-Particle\\Simulations')
                        filename = str(Run_row) + '_' + CarrierGeometry + '_PP-Collision' + '_L' + str(Loading_Replicate) +  '_C' + str(n) + '.rocky'
                        project.SaveProject(filename)
                        user_processes = project.GetUserProcessCollection()
                        particles = study.GetParticles()
                        Contacts = study.GetContactData()
                        ParticleCollection = study.GetParticleCollection()
                        #Physics
                        physics = study.GetPhysics()
                        physics.SetGravityYDirection(float(P_Array_dict['GravityYDirection'][Position]))
                        physics.SetGravityStartTime(float(P_Array_dict['GravityStartTime'][Position]))
                        physics.SetGravityStopTime(float(P_Array_dict['GravityStopTime'][Position]))
                        physics.SetNormalForceModel(str(P_Array_dict['NormalForceModel'][Position]))
                        physics.SetTangentialForceModel(str(P_Array_dict['TangentialForceModel'][Position]))
                        physics.SetAdhesionModel(str(P_Array_dict['AdhesionModel'][Position]))
                        physics.SetRollingResistanceModel(str(P_Array_dict['RollingResistanceModel'][Position]))
                        physics.SetNumericalSofteningFactor(float(P_Array_dict['NumericalSofteningFactor'][Position]))
                        #Modules aktivieren
                        module_collection = study.GetModuleCollection()
                        GetContactsOverlapMonitor = module_collection.GetModule('Contacts Overlap Monitor')
                        GetContactsOverlapMonitor.EnableModule()
                        GetContactsOverlapMonitor.SetModuleProperty('Overlap Warning Level #1', 0.1)
                        GetContactsOverlapMonitor.SetModuleProperty('Overlap Warning Level #2', 2)
                        GetContactsOverlapMonitor.SetModuleProperty('Overlap Warning Level #3', 5)
                        GetInter_particle_Collision_Statistics = module_collection.GetModule('Inter-particle Collision Statistics')
                        GetInter_particle_Collision_Statistics.EnableModule()
                        GetSPHDensityMonitor = module_collection.GetModule('SPH Density Monitor')
                        GetSPHDensityMonitor.DisableModule()
                        GetInter_particle_Collision_Statistics.SetModuleProperty('Duration', True)
                        GetInter_particle_Collision_Statistics.SetModuleProperty('Frequency', True)
                        GetInter_particle_Collision_Statistics.SetModuleProperty('Normal Impact Velocity', True)
                        GetInter_particle_Collision_Statistics.SetModuleProperty('Tangential Impact Velocity', True)
                        GetInter_particle_Collision_Statistics.SetModuleProperty('Power', True)
                        GetInter_particle_Collision_Statistics.SetModuleProperty('Force', True)
                        #Materials
                        MaterialCollection = study.GetMaterialCollection()
                        APIMaterial = MaterialCollection.New()
                        APIMaterial.SetName('API')
                        APIMaterial.SetDensity(float(P_Array_dict['APIDensity'][Position]))
                        APIMaterial.SetYoungsModulus(float(P_Array_dict['APIYoungsModulus'][Position]))
                        APIMaterial.SetPoissonRatio(float(P_Array_dict['APIPoissonRatio'][Position]))
                        CarrierMaterial = MaterialCollection.New()
                        CarrierMaterial.SetName('Carrier')
                        CarrierMaterial.SetDensity(float(P_Array_dict['CarrierDensity'][Position]))
                        CarrierMaterial.SetYoungsModulus(float(P_Array_dict['CarrierYoungsModulus'][Position]))
                        CarrierMaterial.SetPoissonRatio(float(P_Array_dict['CarrierPoissonRatio'][Position]))
                        #Material Interactions
                        material_API = study.GetElement('API')
                        material_Carrier = study.GetElement('Carrier')
                        interaction_collection = study.GetMaterialsInteractionCollection()
                        interaction = interaction_collection.GetMaterialsInteraction(material_API, material_Carrier)
                        interaction.SetSurfaceEnergy(float(P_Array_dict['APICarrierSurfaceEnergy'][Position]))
                        interaction.SetStaticFriction(float(P_Array_dict['APICarrierStaticFriction'][Position]))
                        interaction.SetDynamicFriction(float(P_Array_dict['APICarrierDynamicFriction'][Position]))
                        interaction.SetContactStiffnessMultiplier(float(P_Array_dict['APICarrierContactStiffnessMultiplier'][Position]))
                        interaction.SetRestitutionCoefficient(float(P_Array_dict['APICarrierRestitutionCoefficient'][Position]))
                        interaction = interaction_collection.GetMaterialsInteraction(material_API, material_API)
                        interaction.SetSurfaceEnergy(float(P_Array_dict['APIAPISurfaceEnergy'][Position]))
                        interaction.SetStaticFriction(float(P_Array_dict['APIAPIStaticFriction'][Position]))
                        interaction.SetDynamicFriction(float(P_Array_dict['APIAPIDynamicFriction'][Position]))
                        interaction.SetRestitutionCoefficient(float(P_Array_dict['APIAPIRestitutionCoefficient'][Position]))
                        #Carrier Partikel
                        ParticleCollection = study.GetParticleCollection()
                        CarrierParticle = ParticleCollection.New()
                        CarrierParticle.SetName('CarrierParticle')
                        CarrierParticle.SetMaterial('Carrier')
                        os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Geometrien\Carrier')
                        CarrierParticle.SetShape('custom_polyhedron')
                        CarrierParticle.ImportFromSTL(CarrierGeometry + '.stl', 'as_concave')
                        Carrier_SizeDistribution = CarrierParticle.GetSizeDistributionList()
                        Carrier_SizeDistribution.Clear()
                        CarrierSize = Carrier_SizeDistribution.New()
                        CarrierSize.SetSize(CarrierSizeValue, 'um')
                        CarrierParticle.SetSuperquadricDegree(2.0)
                        CarrierParticle.SetMaterial('Carrier')
                        #API Partikel
                        ParticleCollection = study.GetParticleCollection()
                        APIPartikel = ParticleCollection.New()
                        APIPartikel.SetName('API')
                        APIPartikel.SetMaterial('API')
                        APIPartikel.SetRollingResistance(float(P_Array_dict['APIParticleRollingResistance'][Position]))
                        API_SizeDistribution = APIPartikel.GetSizeDistributionList()
                        API_SizeDistribution.Clear()
                        APISize = API_SizeDistribution.New()
                        APISize.SetSize((float(P_Array_dict['APISize'][Position])), 'um')
                        ParticleCollection = study.GetParticleCollection()
                        #Carrier Input both Partners
                        for Partner in range(1,3):
                            #Carrier Input 
                            os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Particle-Particle\\Custom Input Files')
                            input_collection = study.GetParticleInputCollection()
                            CarrierInput = input_collection.AddCustomInput()
                            CarrierInput.SetName('Carrier Input P' + str(Partner))
                            CarrierInput.SetFilePath(str(Run_row) + '_' + CarrierGeometry + '_Carrier Custom Input File Rocky P' + str(Partner) + '_L' + str(Loading_Replicate) + '_' + 'n' + str(n) + '.xlsx')
                            CarrierInput.SetParticle('CarrierParticle')
                            #API Input
                            API_Input = input_collection.AddCustomInput()
                            API_Input.SetName('API_Input P' + str(Partner))
                            API_Input.SetFilePath(str(Run_row) + '_' + CarrierGeometry + '_API Custom Input File Rocky P' + str(Partner) + '_L' + str(Loading_Replicate) + '_' + 'n' + str(n) + '.xlsx')
                            API_Input.SetParticle(APIPartikel)
                        #Contacts
                        Contacts.EnableCollectContactsData()
                        module_collection = study.GetModuleCollection()
                        GetContactsOverlapMonitor = module_collection.GetModule('Contacts Overlap Monitor')
                        GetContactsOverlapMonitor.DisableModule()
                        GetParticlesEnergySpectra = module_collection.GetModule('Particles Energy Spectra')
                        GetParticlesEnergySpectra.EnableModule()
                        GetParticlesEnergySpectra.SetModuleProperty('Impact Energy', True)
                        #Domain Settings
                        domain_settings = study.GetDomainSettings()
                        domain_settings.DisableUseBoundaryLimits()
                        domain_settings.SetCoordinateLimitsMinValues([float(P_Array_dict['DomainXMin'][Position]), float(P_Array_dict['DomainYMin'][Position]), float(P_Array_dict['DomainZMin'][Position])], 'm')
                        domain_settings.SetCoordinateLimitsMaxValues([float(P_Array_dict['DomainXMax'][Position]), float(P_Array_dict['DomainYMax'][Position]), float(P_Array_dict['DomainZMax'][Position])], 'm')
                        #Solver
                        simulator_run = study.GetSimulatorRun()
                        simulator_run.SetSimulationDuration(Sim_Dur)
                        simulator_run.SetSimulationOutputFrequency(Out_Fre)
                        simulator_run.SetSimulationTarget('CPU')
                        simulator_run.SetNumberOfProcessors(16)
                        simulator_run.SetUseFixedTimestep(True)
                        simulator_run.SetFixedTimestep('3.6e-10')
                        os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Particle-Particle\\Simulations')
                        project.SaveProject()
                        os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
                        wb = openpyxl.load_workbook('Recent.xlsx')
                        Carriersheet = wb['Simulations']
                        Carriersheet['A1'] = filename
                        wb.save('Recent.xlsx')
                        #Simulation durchführen
                        os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Particle-Particle\\Simulations')
                        study.StartSimulation(True, False)
                        #Exit einbauen
                        time.sleep(3)
                        Resume = study.CanResumeSimulation()
                        if Resume == True:
                            FollowUp = False
                            project.CloseProject()
                            break
                        os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
                        Carriersheet['A1'] = 'Empty'
                        wb.save('Recent.xlsx')
                        os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Particle-Particle\\Simulations')
                        project.SaveProject()
                        project.CloseProject()
                        os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
                        wb = openpyxl.load_workbook('Carrier.xlsx')
                        Carriersheet = wb['Carrier']
                        cell = Carriersheet['D' + str(Excel_row)]
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        Carrier_Status = str(Loading_Replicate) + '.' + str(n)
                        Carriersheet['D' + str(Excel_row)] = Carrier_Status
                        wb.save('Carrier.xlsx')
            if Decision == 'no':
                project.CloseProject(False)

def ParticleWall_Create():
    global Array_dict
    global P_Array_dict
    global Run_Opt
    global Range
    global FollowUp
    #Define Process
    Process = 'ParticleWall'
    for ParaPosition in range(1,P_Array_lenght + 2):
        if P_Array_dict['Parameter'][ParaPosition - 1] == Process:
            Position = ParaPosition - 1
            break
    global Simulation_Letter
    Simulation_Letter = 'F'
    Span()
    for Excel_row in Range:
        Array_row = Excel_row - 2
        Run_row = Excel_row - 1
        os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
        wb = openpyxl.load_workbook('Carrier.xlsx')
        Carriersheet = wb['Carrier']
        Collision_Maximum = str(Loading_Replicas) + '.' + str(Replicas)
        if Carriersheet['C' + str(Excel_row)].value != None and Carriersheet['F' + str(Excel_row)].value != Collision_Maximum:
            #Finished Simulation Check
            CarrierGeometry = str((Array_dict['Geometry'][Array_row]))
            for Loading_Replicate in range(1,Loading_Replicas+1):
                os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Loading\\Simulations')
                if FollowUp == False:
                    break
                app.OpenProject(str(Run_row) + '_' + CarrierGeometry + '_L' + str(Loading_Replicate) + '_Loading.rocky') 
                study = api.GetStudy()
                Finished = study.HasResults()
                project = api.GetProject()
                if Finished == False:
                    Decision = 'no'
                elif Finished == True:
                    Decision = 'yes'
                if Decision == 'yes':
                    #Define Parameters
                    for ParaPosition in range(1,P_Array_lenght + 2):
                        if P_Array_dict['Parameter'][ParaPosition - 1] == Process:
                            Position = ParaPosition - 1
                            break
                    for Parameter in P_Array_dict:
                        Value = P_Array_dict[Parameter][Position]
                        if Value == 'DoE':
                            DoE_Array_Value = Array_dict[Parameter]
                            DoE_Value = DoE_Array_Value[Array_row]
                            P_Array_dict[Parameter][Position] = DoE_Value
                    CarrierSizeValue = float(P_Array_dict['CarrierParticleSize'][Position]) * 10**(-6)
                    Sim_Dur = float(P_Array_dict['SimulationDuration'][Position])
                    Out_Fre = float(P_Array_dict['SimulationOutputFrequency'][Position])
                    #Projekt
                    project = app.GetProject()
                    study = project.GetStudy()
                    user_processes = project.GetUserProcessCollection()
                    particles = study.GetParticles()
                    Contacts = study.GetContactData()
                    #Carrier
                    CarrierPosition = user_processes.CreatePropertyProcess(particles, 'Carrier')
                    CarrierPosition.SetPropertyGridFunction('Particle Group')
                    CarrierPosition.SetCutValue(0)
                    ##Direkte Kontakte
                    #Carrier
                    CarrierPosition = user_processes.CreatePropertyProcess(particles, 'CarrierPosition')
                    CarrierPosition.SetPropertyGridFunction('Particle Group')
                    CarrierPosition.SetCutValue(0)
                    #Velocity API
                    API_Particles = user_processes.CreatePropertyProcess(particles, 'API_Particles')
                    API_Particles.SetPropertyGridFunction('Particle Group')
                    API_Particles.SetCutValue(1)
                    API_0_V = user_processes.CreatePropertyProcess(API_Particles, '0_Velocity')
                    API_0_V.SetPropertyGridFunction('Absolute Translational Velocity')
                    API_0_V.SetType('Range')
                    API_0_V.SetMaxValue(float(P_Array_dict['APIVelocity'][Position]))
                    #Create Cube
                    Get_API = user_processes.GetProcess('0_Velocity')
                    Cube = user_processes.CreateCubeProcess(Get_API, 'Cube')
                    Cube.SetCenter(0.0, 0.0, 0.0, 'm')
                    Cube.SetSize(20.0, 10.0, 10.0, 'mm')
                    #Create Eulerian
                    Eulerian = user_processes.CreateEulerianStatistics(Cube, 'Eulerian')
                    Eulerian.SetDivisions((1, 1, 1,))
                    #Time Step
                    os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Loading\\Results')
                    wb = openpyxl.load_workbook('Results_Loading.xlsx')
                    sheet = wb['Total']
                    #Domain Bounds
                    Process = 'Loading'
                    for ParaPosition in range(1,P_Array_lenght + 2):
                        if P_Array_dict['Parameter'][ParaPosition - 1] == Process:
                            Position = ParaPosition - 1
                            break
                    DomainBounds = float(P_Array_dict['DomainXMax'][Position])
                    SimulationDuration = float(P_Array_dict['SimulationDuration'][Position])
                    OutputFrequency = float(P_Array_dict['SimulationOutputFrequency'][Position])
                    Process = 'ParticleWall'
                    for ParaPosition in range(1,P_Array_lenght + 2):
                        if P_Array_dict['Parameter'][ParaPosition - 1] == Process:
                            Position = ParaPosition - 1
                            break
                    TS = int(SimulationDuration / OutputFrequency)
                    Carrier_Particle_V = user_processes.CreatePropertyProcess(particles, 'Carrier_Particle_V')
                    Carrier_Particle_V.SetPropertyGridFunction('Particle Group')
                    Carrier_Particle_V.SetCutValue(0)
                    Carrier_V = float(Carrier_Particle_V.GetGridFunction('Absolute Translational Velocity').GetArray(time_step=TS)[0])
                    API_Resting_V = Carrier_V * 2
                    API_0_V.SetMaxValue(API_Resting_V)
                    for n in range(1,n_max):
                        #Write New input files
                        #Carrier
                        Carrier_Angle = float(CarrierPosition.GetGridFunction('Orientation Angle').GetArray(time_step=TS))
                        Carrier_X_Orientation = float(CarrierPosition.GetGridFunction('Orientation Vector X').GetArray(time_step=TS))
                        Carrier_Y_Orientation = float(CarrierPosition.GetGridFunction('Orientation Vector Y').GetArray(time_step=TS))
                        Carrier_Z_Orientation = float(CarrierPosition.GetGridFunction('Orientation Vector Z').GetArray(time_step=TS))
                        os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Particle-Wall\\Custom Input Files')
                        wb = openpyxl.Workbook()
                        sheet = wb.active
                        sheet.title = 'Carrier'  
                        filename = (str(Run_row) + '_' + CarrierGeometry + '_Carrier Custom Input File Rocky_L' + str(Loading_Replicate) + '_' + 'n' + str(n) + '.xlsx')
                        wb.save(filename)
                        wb = openpyxl.load_workbook(filename)
                        sheet = wb['Carrier'] 
                        sheet['A1'] = 'x'
                        sheet['B1'] = 'y'
                        sheet['C1'] = 'z'
                        sheet['A2'] = 0
                        sheet['B2'] = 0
                        sheet['C2'] = 0
                        sheet['D1'] = 'size'
                        sheet['D2'] = CarrierSizeValue
                        sheet['F1'] = 'nx'
                        sheet['G1'] = 'ny'
                        sheet['H1'] = 'nz'
                        sheet['I1'] = 'ux'
                        sheet['J1'] = 'uy'
                        sheet['K1'] = 'uz'
                        sheet['I2'] = 0
                        sheet['J2'] = (float(P_Array_dict['CollisionVelocity'][Position]))
                        sheet['K2'] = 0
                        Random_Angle = random.randint(0,90)
                        rad_Angle = Random_Angle * (3.14159265359 / 180)
                        x_rotation = np.random.rand()
                        y_rotation = np.random.rand()
                        z_rotation = np.random.rand()
                        orientation_vector = np.array([x_rotation, y_rotation, z_rotation])
                        orientation_vector /= np.linalg.norm(orientation_vector)
                        sheet['E1'] = 'angle'
                        sheet['E2'] = rad_Angle
                        sheet['F2'] = orientation_vector[0]
                        sheet['G2'] = orientation_vector[1]
                        sheet['H2'] = orientation_vector[2]
                        wb.save(filename)
                        CarrierPosition_X = float(CarrierPosition.GetGridFunction('Particle X-Coordinate').GetArray(time_step=TS))
                        CarrierPosition_Y = float(CarrierPosition.GetGridFunction('Particle Y-Coordinate').GetArray(time_step=TS))
                        CarrierPosition_Z = float(CarrierPosition.GetGridFunction('Particle Z-Coordinate').GetArray(time_step=TS))
                        #API Position
                        API_X_array = API_0_V.GetGridFunction('Particle X-Coordinate').GetArray(time_step = TS)
                        API_Y_array = API_0_V.GetGridFunction('Particle Y-Coordinate').GetArray(time_step = TS)
                        API_Z_array = API_0_V.GetGridFunction('Particle Z-Coordinate').GetArray(time_step = TS)
                        #API
                        os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Particle-Wall\\Custom Input Files')
                        wb = openpyxl.Workbook()
                        sheet = wb.active
                        sheet.title = 'API'
                        filename = (str(Run_row) + '_' + CarrierGeometry + '_API Custom Input File Rocky_L' + str(Loading_Replicate) + '_' + 'n' + str(n) + '.xlsx')
                        wb.save(filename)
                        wb = openpyxl.load_workbook(filename)
                        API_sheet = wb['API']
                        API_sheet['A1'] = 'x'
                        API_sheet['B1'] = 'y'
                        API_sheet['C1'] = 'z'
                        API_sheet['D1'] = 'size'
                        API_sheet['E1'] = 'ux'
                        API_sheet['F1'] = 'uy'
                        API_sheet['G1'] = 'uz'
                        numrows_1 = len(API_X_array)
                        Distance = 2 * DomainBounds
                        for a in range(0, numrows_1):
                            API_Diameter = float(API_0_V.GetGridFunction('Particle Size').GetArray(time_step=TS)[a])
                            #Shift Correction
                            #X
                            if (API_X_array[a] - CarrierPosition_X) > CarrierSizeValue:
                                API_X_array[a] = API_X_array[a] - Distance
                            elif (API_X_array[a] - CarrierPosition_X) < - CarrierSizeValue:
                                API_X_array[a] = API_X_array[a] + Distance
                            #Y
                            if (API_Y_array[a] - CarrierPosition_Y) > CarrierSizeValue:
                                API_Y_array[a] = API_Y_array[a] - Distance
                            elif (API_Y_array[a] - CarrierPosition_Y) < - CarrierSizeValue:
                                API_Y_array[a] = API_Y_array[a] + Distance
                            #Z
                            if (API_Z_array[a] - CarrierPosition_Z) > CarrierSizeValue:
                                API_Z_array[a] = API_Z_array[a] - Distance
                            elif (API_Z_array[a] - CarrierPosition_Z) < - CarrierSizeValue:
                                API_Z_array[a] = API_Z_array[a] + Distance
                            #Position
                            API_X_Shift = float(API_X_array[a]) - CarrierPosition_X
                            API_Y_Shift = float(API_Y_array[a]) - CarrierPosition_Y
                            API_Z_Shift = float(API_Z_array[a]) - CarrierPosition_Z
                            #First Rotation
                            angle = Carrier_Angle
                            x_orient = Carrier_X_Orientation
                            y_orient = Carrier_Y_Orientation
                            z_orient = Carrier_Z_Orientation
                            P = np.array([API_X_Shift, API_Y_Shift, API_Z_Shift])
                            rotation_matrix = np.array([
                            [np.cos(angle) + x_orient**2 * (1 - np.cos(angle)),
                             x_orient * y_orient * (1 - np.cos(angle)) - z_orient * np.sin(angle),
                             x_orient * z_orient * (1 - np.cos(angle)) + y_orient * np.sin(angle)],
                            [y_orient * x_orient * (1 - np.cos(angle)) + z_orient * np.sin(angle),
                             np.cos(angle) + y_orient**2 * (1 - np.cos(angle)),
                             y_orient * z_orient * (1 - np.cos(angle)) - x_orient * np.sin(angle)],
                            [z_orient * x_orient * (1 - np.cos(angle)) - y_orient * np.sin(angle),
                             z_orient * y_orient * (1 - np.cos(angle)) + x_orient * np.sin(angle),
                             np.cos(angle) + z_orient**2 * (1 - np.cos(angle))]])
                            inverse_rotation_matrix = np.linalg.inv(rotation_matrix)
                            rotated_1_P = np.dot(inverse_rotation_matrix, P)
                            #Second Rotation
                            angle = rad_Angle
                            x_orient = orientation_vector[0]
                            y_orient = orientation_vector[1]
                            z_orient = orientation_vector[2]
                            P = rotated_1_P
                            rotation_matrix = np.array([
                            [np.cos(angle) + x_orient**2 * (1 - np.cos(angle)),
                             x_orient * y_orient * (1 - np.cos(angle)) - z_orient * np.sin(angle),
                             x_orient * z_orient * (1 - np.cos(angle)) + y_orient * np.sin(angle)],
                            [y_orient * x_orient * (1 - np.cos(angle)) + z_orient * np.sin(angle),
                             np.cos(angle) + y_orient**2 * (1 - np.cos(angle)),
                             y_orient * z_orient * (1 - np.cos(angle)) - x_orient * np.sin(angle)],
                            [z_orient * x_orient * (1 - np.cos(angle)) - y_orient * np.sin(angle),
                             z_orient * y_orient * (1 - np.cos(angle)) + x_orient * np.sin(angle),
                             np.cos(angle) + z_orient**2 * (1 - np.cos(angle))]])
                            rotated_P = np.dot(rotation_matrix, P)
                            #Delete Outer Particles
                            X_Position = rotated_P[0]
                            Y_Position = rotated_P[1]
                            Z_Position = rotated_P[2]
                            #X
                            if X_Position > CarrierSizeValue * 1.1 or X_Position < - CarrierSizeValue* 1.1:
                                X_Position = (DomainBounds / 2) * 0.9
                                Y_Position = (DomainBounds / 2) * 0.9
                                Z_Position = (DomainBounds / 2) * 0.9
                            #Y
                            if Y_Position > CarrierSizeValue* 1.1 or Y_Position < - CarrierSizeValue* 1.1:
                                X_Position = (DomainBounds / 2) * 0.9
                                Y_Position = (DomainBounds / 2) * 0.9
                                Z_Position = (DomainBounds / 2) * 0.9
                            #Z
                            if Z_Position > CarrierSizeValue* 1.1 or Z_Position < - CarrierSizeValue* 1.1:
                                X_Position = (DomainBounds / 2) * 0.9
                                Y_Position = (DomainBounds / 2) * 0.9
                                Z_Position = (DomainBounds / 2) * 0.9
                            #Eintrag
                            position_factor = 1.0
                            API_sheet['A' + str(a + 2)] = X_Position * position_factor
                            API_sheet['B' + str(a + 2)] = Y_Position * position_factor
                            API_sheet['C' + str(a + 2)] = Z_Position * position_factor
                            API_sheet['D' + str(a + 2)] = API_Diameter
                            API_sheet['E' + str(a + 2)] = 0
                            API_sheet['F' + str(a + 2)] = (float(P_Array_dict['CollisionVelocity'][Position]))
                            API_sheet['G' + str(a + 2)] = 0
                        wb.save(filename)
                    project.CloseProject(False)
                for n in range(1,n_max):
                    os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
                    wb = openpyxl.load_workbook('Carrier.xlsx')
                    Carriersheet = wb['Carrier']
                    Finished_n = Carriersheet['F' + str(Excel_row)].value
                    Collision_Maximum = str(Loading_Replicas) + '.' + str(Replicas)
                    if Finished_n == None or Finished_n != Collision_Maximum:
                        #Simulation erstellen
                        app.CreateProject()
                        project = app.GetProject()
                        study = project.GetStudy()
                        study.SetCustomerName(CarrierGeometry)
                        os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Particle-Wall\\Simulations')
                        filename = str(Run_row) + '_' + CarrierGeometry + '_PW-Collision' + '_L' + str(Loading_Replicate) +  '_C' + str(n) + '.rocky'
                        project.SaveProject(filename)                        
                        user_processes = project.GetUserProcessCollection()
                        particles = study.GetParticles()
                        Contacts = study.GetContactData()
                        ParticleCollection = study.GetParticleCollection()
                        #Physics
                        physics = study.GetPhysics()
                        physics.SetGravityYDirection(float(P_Array_dict['GravityYDirection'][Position]))
                        physics.SetGravityStartTime(float(P_Array_dict['GravityStartTime'][Position]))
                        physics.SetGravityStopTime(float(P_Array_dict['GravityStopTime'][Position]))
                        physics.SetNormalForceModel(str(P_Array_dict['NormalForceModel'][Position]))
                        physics.SetTangentialForceModel(str(P_Array_dict['TangentialForceModel'][Position]))
                        physics.SetAdhesionModel(str(P_Array_dict['AdhesionModel'][Position]))
                        physics.SetRollingResistanceModel(str(P_Array_dict['RollingResistanceModel'][Position]))
                        physics.SetNumericalSofteningFactor(float(P_Array_dict['NumericalSofteningFactor'][Position]))
                        #Modules aktivieren
                        module_collection = study.GetModuleCollection()
                        GetContactsOverlapMonitor = module_collection.GetModule('Contacts Overlap Monitor')
                        GetContactsOverlapMonitor.EnableModule()
                        GetContactsOverlapMonitor.SetModuleProperty('Overlap Warning Level #1', 0.1)
                        GetContactsOverlapMonitor.SetModuleProperty('Overlap Warning Level #2', 2)
                        GetContactsOverlapMonitor.SetModuleProperty('Overlap Warning Level #3', 5)
                        GetInter_particle_Collision_Statistics = module_collection.GetModule('Inter-particle Collision Statistics')
                        GetInter_particle_Collision_Statistics.EnableModule()
                        GetSPHDensityMonitor = module_collection.GetModule('SPH Density Monitor')
                        GetSPHDensityMonitor.DisableModule()
                        GetInter_particle_Collision_Statistics.SetModuleProperty('Duration', True)
                        GetInter_particle_Collision_Statistics.SetModuleProperty('Frequency', True)
                        GetInter_particle_Collision_Statistics.SetModuleProperty('Normal Impact Velocity', True)
                        GetInter_particle_Collision_Statistics.SetModuleProperty('Tangential Impact Velocity', True)
                        GetInter_particle_Collision_Statistics.SetModuleProperty('Power', True)
                        GetInter_particle_Collision_Statistics.SetModuleProperty('Force', True)
                        #Materials
                        MaterialCollection = study.GetMaterialCollection()
                        APIMaterial = MaterialCollection.New()
                        APIMaterial.SetName('API')
                        APIMaterial.SetDensity(float(P_Array_dict['APIDensity'][Position]))
                        APIMaterial.SetYoungsModulus(float(P_Array_dict['APIYoungsModulus'][Position]))
                        APIMaterial.SetPoissonRatio(float(P_Array_dict['APIPoissonRatio'][Position]))
                        CarrierMaterial = MaterialCollection.New()
                        CarrierMaterial.SetName('Carrier')
                        CarrierMaterial.SetDensity(float(P_Array_dict['CarrierDensity'][Position]))
                        CarrierMaterial.SetYoungsModulus(float(P_Array_dict['CarrierYoungsModulus'][Position]))
                        CarrierMaterial.SetPoissonRatio(float(P_Array_dict['CarrierPoissonRatio'][Position]))
                        #Material Interactions
                        material_API = study.GetElement('API')
                        material_Carrier = study.GetElement('Carrier')
                        interaction_collection = study.GetMaterialsInteractionCollection()
                        interaction = interaction_collection.GetMaterialsInteraction(material_API, material_Carrier)
                        interaction.SetSurfaceEnergy(float(P_Array_dict['APICarrierSurfaceEnergy'][Position]))
                        interaction.SetStaticFriction(float(P_Array_dict['APICarrierStaticFriction'][Position]))
                        interaction.SetDynamicFriction(float(P_Array_dict['APICarrierDynamicFriction'][Position]))
                        interaction.SetContactStiffnessMultiplier(float(P_Array_dict['APICarrierContactStiffnessMultiplier'][Position]))
                        interaction.SetRestitutionCoefficient(float(P_Array_dict['APICarrierRestitutionCoefficient'][Position]))
                        interaction = interaction_collection.GetMaterialsInteraction(material_API, material_API)
                        interaction.SetSurfaceEnergy(float(P_Array_dict['APIAPISurfaceEnergy'][Position]))
                        interaction.SetStaticFriction(float(P_Array_dict['APIAPIStaticFriction'][Position]))
                        interaction.SetDynamicFriction(float(P_Array_dict['APIAPIDynamicFriction'][Position]))
                        interaction.SetRestitutionCoefficient(float(P_Array_dict['APIAPIRestitutionCoefficient'][Position]))
                        #Platte
                        geometry_collection = study.GetGeometryCollection()
                        os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Geometrien\\Platten')
                        Angle = (int(P_Array_dict['CollisionAngle'][Position]))
                        study.ImportWall('Platte' + str(Angle) + '.stl', 1.0, False)
                        Platte = geometry_collection.GetGeometry('Platte' + str(Angle))
                        Platte.SetVerticalOffset(float(P_Array_dict['PlatteVerticalOffset'][Position]))
                        #Carrier Partikel
                        ParticleCollection = study.GetParticleCollection()
                        CarrierParticle = ParticleCollection.New()
                        CarrierParticle.SetName('CarrierParticle')
                        os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Geometrien\Carrier')
                        CarrierParticle.SetShape('custom_polyhedron')
                        CarrierParticle.ImportFromSTL(CarrierGeometry + '.stl', 'as_concave')
                        Carrier_SizeDistribution = CarrierParticle.GetSizeDistributionList()
                        Carrier_SizeDistribution.Clear()
                        CarrierSize = Carrier_SizeDistribution.New()
                        CarrierSize.SetSize(CarrierSizeValue, 'um')
                        CarrierParticle.SetSuperquadricDegree(2.0)
                        CarrierParticle.SetMaterial('Carrier')
                        #API Partikel
                        ParticleCollection = study.GetParticleCollection()
                        APIPartikel = ParticleCollection.New()
                        APIPartikel.SetName('API')
                        APIPartikel.SetMaterial('API')
                        APIPartikel.SetRollingResistance(float(P_Array_dict['APIParticleRollingResistance'][Position]))
                        API_SizeDistribution = APIPartikel.GetSizeDistributionList()
                        API_SizeDistribution.Clear()
                        APISize = API_SizeDistribution.New()
                        APISize.SetSize((float(P_Array_dict['APISize'][Position])), 'um')
                        ParticleCollection = study.GetParticleCollection()
                        #Carrier Input 
                        os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Particle-Wall\\Custom Input Files')
                        input_collection = study.GetParticleInputCollection()
                        CarrierInput = input_collection.AddCustomInput()
                        CarrierInput.SetName('Carrier Input')
                        filename = (str(Run_row) + '_' + CarrierGeometry + '_Carrier Custom Input File Rocky_L' + str(Loading_Replicate) + '_' + 'n' + str(n) + '.xlsx')
                        CarrierInput.SetFilePath(filename)
                        CarrierInput.SetParticle('CarrierParticle')
                        #API Input
                        API_Input = input_collection.AddCustomInput()
                        API_Input.SetName('API_Input')
                        filename = (str(Run_row) + '_' + CarrierGeometry + '_API Custom Input File Rocky_L' + str(Loading_Replicate) + '_' + 'n' + str(n) + '.xlsx')
                        API_Input.SetFilePath(filename)
                        API_Input.SetParticle(APIPartikel)
                        #Contacts
                        Contacts.EnableCollectContactsData()
                        module_collection = study.GetModuleCollection()
                        GetParticlesEnergySpectra = module_collection.GetModule('Particles Energy Spectra')
                        GetParticlesEnergySpectra.EnableModule()
                        GetParticlesEnergySpectra.SetModuleProperty('Impact Energy', True)
                        #Domain Settings
                        domain_settings = study.GetDomainSettings()
                        domain_settings.DisableUseBoundaryLimits()
                        domain_settings.SetCoordinateLimitsMinValues([float(P_Array_dict['DomainXMin'][Position]), float(P_Array_dict['DomainYMin'][Position]), float(P_Array_dict['DomainZMin'][Position])], 'm')
                        domain_settings.SetCoordinateLimitsMaxValues([float(P_Array_dict['DomainXMax'][Position]), float(P_Array_dict['DomainYMax'][Position]), float(P_Array_dict['DomainZMax'][Position])], 'm')
                        #Solver
                        simulator_run = study.GetSimulatorRun()
                        simulator_run.SetSimulationDuration(Sim_Dur)
                        simulator_run.SetSimulationOutputFrequency(Out_Fre)
                        simulator_run.SetSimulationTarget('CPU')
                        simulator_run.SetNumberOfProcessors(16)
                        simulator_run.SetUseFixedTimestep(True)
                        simulator_run.SetFixedTimestep('3.6e-10')
                        os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Particle-Wall\\Simulations')
                        project.SaveProject()
                        os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
                        wb = openpyxl.load_workbook('Recent.xlsx')
                        Carriersheet = wb['Simulations']
                        filename = str(Run_row) + '_' + CarrierGeometry + '_PW-Collision' + '_L' + str(Loading_Replicate) +  '_C' + str(n) + '.rocky'
                        Carriersheet['A1'] = filename
                        wb.save('Recent.xlsx')
                        #Simulation durchführen
                        os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Particle-Wall\\Simulations')
                        study.StartSimulation(True, False)
                        #Exit einbauen
                        time.sleep(3)
                        Resume = study.CanResumeSimulation()
                        if Resume == True:
                            FollowUp = False
                            project.CloseProject()
                            break
                        os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
                        Carriersheet['A1'] = 'Empty'
                        wb.save('Recent.xlsx')
                        os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Particle-Wall\\Simulations')
                        project.SaveProject()
                        project.CloseProject()
                        os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
                        wb = openpyxl.load_workbook('Carrier.xlsx')
                        Carriersheet = wb['Carrier']
                        cell = Carriersheet['F' + str(Excel_row)]
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        Carrier_Status = str(Loading_Replicate) + '.' + str(n)
                        Carriersheet['F' + str(Excel_row)] = Carrier_Status
                        wb.save('Carrier.xlsx')
            if Decision == 'no':
                project.CloseProject(False)

def ParticleParticle_Analyse():
    global Array_dict
    global P_Array_dict
    global Run_Opt
    global Range
    #PostParticle-Particle
    Process = 'ParticleParticle'
    for ParaPosition in range(1,P_Array_lenght + 2):
        if P_Array_dict['Parameter'][ParaPosition - 1] == Process:
            Position = ParaPosition - 1
            break
    os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\\Particle-Particle\\Results')
    if not os.path.exists('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\\Particle-Particle\\Results\\Results_Particle-Particle.xlsx'):
        wb = openpyxl.Workbook()
        wb.save('Results_Particle-Particle.xlsx')
    os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\\Particle-Particle\\Results')
    wb = openpyxl.load_workbook('Results_Particle-Particle.xlsx')
    sheet = wb.active
    sheet.title = 'Raw'
    sheet = wb['Raw']
    sheet['A1'] = 'No'
    sheet['B1'] = 'PreAPI'
    sheet['C1'] = 'PostAPI'
    sheet['D1'] = 'PreAPI mass [ng]'
    sheet['E1'] = 'PostAPI mass [ng]'
    sheet['F1'] = 'absolute Detachment [ng]'
    sheet['G1'] = 'relative Detachment [%]'
    sheet['H1'] = 'RSD [%]'
    sheet['I1'] = 'CRD'
    sheet['J1'] = 'Last TS'
    wb.save('Results_Particle-Particle.xlsx')
    Individual_Name = 'Individual'
    if Individual_Name not in wb.sheetnames:
        wb.create_sheet('Individual')
        Indi_sheet = wb['Individual']
    else:
        Indi_sheet = wb['Individual']
    wb.save('Results_Particle-Particle.xlsx')
    global Simulation_Letter
    Simulation_Letter = 'E'
    Span()
    for Excel_row in Range:
        Array_row = Excel_row - 2
        Run_row = Excel_row - 1
        os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
        wb = openpyxl.load_workbook('Carrier.xlsx')
        Carriersheet = wb['Carrier']
        if Carriersheet['D' + str(Excel_row)].value != None and Carriersheet['E' + str(Excel_row)].value == None:
            CarrierGeometry = str(Carriersheet['A' + str(Excel_row)].value)
            #Define Parameters
            for Parameter in P_Array_dict:
                Value = P_Array_dict[Parameter][Position]
                if Value == 'DoE':
                    DoE_Array_Value = Array_dict[Parameter]
                    DoE_Value = DoE_Array_Value[Array_row]
                    P_Array_dict[Parameter][Position] = DoE_Value
            PreAPI_Array = []
            PostAPI_Array = []
            PreAPI_mass_Array = []
            PostAPI_mass_Array = []
            absolute_Detachment_Array = []
            relative_Detachment_Array = []
            CRD_Array = []
            Last_TS_Array = []
            #Finished Simulation Check
            for Loading_Replicate in range(1,Loading_Replicas+1):
                os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Results\Particle-Particle\Results')
                wb = openpyxl.load_workbook('Results_Particle-Particle.xlsx')
                Indi_sheet = wb['Individual']
                Indi_Array = []
                #Indi_Row
                for row in range(2,Indi_sheet.max_row+2):
                    if Indi_sheet['A' + str(row)].value == None:
                        Indi_row = row
                        break
                for n in range(1,n_max):
                    filepath_PPSimulations = 'D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Particle-Particle\\Simulations'
                    os.chdir(filepath_PPSimulations)
                    filename_PPSimulation = str(Run_row) + '_' + CarrierGeometry + '_PP-Collision' + '_L' + str(Loading_Replicate) +  '_C' + str(n) + '.rocky'
                    app.OpenProject(filename_PPSimulation)        
                    study = api.GetStudy()
                    Finished = study.HasResults()
                    Resume = study.CanResumeSimulation()
                    project = api.GetProject()
                    if Finished == False:
                        Decision = 'no'
                    elif Finished == True:
                        Decision = 'yes'
                    if Resume == True:
                        Decision = 'no'
                    #Bei Entscheidung Ja
                    if Decision == 'yes':
                        #Projekt
                        Pre_TS = 1
                        project = app.GetProject()
                        study = api.GetStudy()
                        user_processes = project.GetUserProcessCollection()
                        particles = study.GetParticles()
                        Contacts = study.GetContactData()
                        #Carrier
                        CarrierPosition = user_processes.CreatePropertyProcess(particles, 'CarrierPosition')
                        CarrierPosition.SetPropertyGridFunction('Particle Group')
                        CarrierPosition.SetCutValue(0)
                        #Create Cube
                        Get_Carrier_Process = user_processes.GetProcess('CarrierPosition')
                        Cube_Carrier = user_processes.CreateCubeProcess(Get_Carrier_Process, 'Cube_Carrier')
                        Cube_Carrier.SetCenter(0.0, 0.0, 0.0, 'm')
                        Cube_Carrier.SetSize(20.0, 10.0, 10.0, 'mm')
                        #Create Eulerian
                        Eulerian_Carrier = user_processes.CreateEulerianStatistics(Cube_Carrier, 'Eulerian_Carrier')
                        Eulerian_Carrier.SetDivisions((1, 1, 1,))
                        #Contact from Carrier
                        From_Carrier = user_processes.CreateContactToParticleProcess(Contacts, 'To Carrier')
                        #API
                        #Contact to Particle
                        Contact_to_1 = user_processes.CreatePropertyProcess(From_Carrier, 'Contact_to_1')
                        Contact_to_1.SetPropertyGridFunction('Particle Group')
                        Contact_to_1.SetCutValue(1)
                        #Create Cube
                        Get_Contact_to_1 = user_processes.GetProcess('Contact_to_1')
                        Cube_1 = user_processes.CreateCubeProcess(Get_Contact_to_1, 'Cube_1')
                        Cube_1.SetCenter(0.0, 0.0, 0.0, 'm')
                        Cube_1.SetSize(20.0, 10.0, 10.0, 'mm')
                        #Create Eulerian
                        Eulerian_1 = user_processes.CreateEulerianStatistics(Cube_1, 'Eulerian_1')
                        Eulerian_1.SetDivisions((1, 1, 1,))
                        #API to API
                        API_to_API = user_processes.CreateParticleToContactProcess(Contact_to_1, 'From_API')
                        From_API_to_API = user_processes.CreatePropertyProcess(API_to_API, 'From_API_to_API')
                        From_API_to_API.SetPropertyGridFunction('Particle Group From')
                        From_API_to_API.SetCutValue(1)
                        contact_API_API = user_processes.CreateContactToParticleProcess(From_API_to_API, 'contact_API_API')
                        Contact_API_API = user_processes.CreatePropertyProcess(contact_API_API, 'Contact_API_API')
                        Contact_API_API.SetPropertyGridFunction('Particle Group')
                        Contact_API_API.SetCutValue(1)
                        #Create Cube
                        Get_Contact_to_API = user_processes.GetProcess('Contact_API_API')
                        Cube_API = user_processes.CreateCubeProcess(Get_Contact_to_API, 'Cube_API')
                        Cube_API.SetCenter(0.0, 0.0, 0.0, 'm')
                        Cube_API.SetSize(20.0, 10.0, 10.0, 'mm')
                        #Create Eulerian
                        Eulerian_API = user_processes.CreateEulerianStatistics(Cube_API, 'Eulerian_API')
                        Eulerian_API.SetDivisions((1, 1, 1,))
                        #Write Results
                        Particle_Mass_Process = user_processes.CreatePropertyProcess(particles)
                        Particle_Mass_Process.SetName('ParticleMassProcess')
                        Particle_Mass_Process.SetPropertyGridFunction('Particle Group')
                        Particle_Mass_Process.SetCutValue(0)
                        Carrier_Mass = float(Particle_Mass_Process.GetGridFunction('Particle Mass').GetArray(time_step=1)[0]) * 10e11
                        Particle_Mass_Process.SetCutValue(1)
                        API_Mass = float(Particle_Mass_Process.GetGridFunction('Particle Mass').GetArray(time_step=1)[0]) * 10e11
                        #API in Kontakt und 2. Schicht zusammenführen
                        X_API = Contact_API_API.GetGridFunction('Particle X-Coordinate').GetArray(time_step = Pre_TS)
                        Y_API = Contact_API_API.GetGridFunction('Particle Y-Coordinate').GetArray(time_step = Pre_TS)
                        Z_API = Contact_API_API.GetGridFunction('Particle Z-Coordinate').GetArray(time_step = Pre_TS)
                        Matrix_Contact_API = np.array([X_API, Y_API, Z_API])
                        X_Carrier = Contact_to_1.GetGridFunction('Particle X-Coordinate').GetArray(time_step = Pre_TS)
                        Y_Carrier = Contact_to_1.GetGridFunction('Particle Y-Coordinate').GetArray(time_step = Pre_TS)
                        Z_Carrier = Contact_to_1.GetGridFunction('Particle Z-Coordinate').GetArray(time_step = Pre_TS)
                        Matrix_Contact_Carrier = np.array([X_Carrier, Y_Carrier, Z_Carrier])
                        Matrix_Carrier = Matrix_Contact_Carrier.T
                        Matrix_API = Matrix_Contact_API.T
                        def row_in_matrix(row, matrix):
                            if matrix.size == 0:
                                return False
                            return np.any(np.all(row == matrix, axis=1))
                        Total_API = []
                        if not Matrix_API.size == 0:
                            for i, row in enumerate(Matrix_API):
                                if not row_in_matrix(row, Matrix_Carrier):
                                    Total_API.append(i)
                        PreAPI = len(X_Carrier) + len(Total_API)
                        #PreAPI = len(Total_API)
                        Particle_Mass_Process.SetCutValue(1)
                        API_Array = []
                        for mass in range(0,PreAPI):
                            API_Mass = float(Particle_Mass_Process.GetGridFunction('Particle Mass').GetArray(time_step=Pre_TS)[mass]) * 10e11
                            API_Array.append(API_Mass)
                        PreAPI_mass = sum(API_Array)
                        SimulatorRun = study.GetSimulatorRun()
                        OutFre = SimulatorRun.GetOutputFrequency()
                        SimDur = SimulatorRun.GetSimulationDuration()
                        Last_TS = int(SimDur / OutFre)
                        #Individual Input
                        #Last_TS = 
                        a = float(Eulerian_1.GetGridFunction('Number of Particles').GetArray(time_step = 3))
                        Post_TS = Last_TS
                        for TS in range(4,Last_TS):
                            b = float(Eulerian_1.GetGridFunction('Number of Particles').GetArray(time_step = TS))
                            if b < a:
                                Post_TS = TS
                                break
                            a = b
                        if Post_TS != Last_TS:
                            a = float(Eulerian_1.GetGridFunction('Number of Particles').GetArray(time_step = Post_TS + 1))
                            for TS in range(Post_TS + 2, Last_TS):
                                b = float(Eulerian_1.GetGridFunction('Number of Particles').GetArray(time_step = TS))
                                if b >= a:
                                    Post_TS = TS + 5
                                    break
                                a = b
                        if Post_TS > Last_TS:
                            Post_TS = Last_TS
                        Carrier_Elimination = 100
                        for this_time in range(1,Last_TS):
                            if int(Eulerian_Carrier.GetGridFunction('Number of Particles').GetArray(time_step = this_time)) == 0:
                                Carrier_Elimination = this_time
                                break
                        if Post_TS > Carrier_Elimination:
                            Post_TS = Carrier_Elimination - 5
                        #Individual Post_TS
                        Post_TS = 30
                        #API in Kontakt und 2. Schicht zusammenführen
                        X_API = Contact_API_API.GetGridFunction('Particle X-Coordinate').GetArray(time_step = Post_TS)
                        Y_API = Contact_API_API.GetGridFunction('Particle Y-Coordinate').GetArray(time_step = Post_TS)
                        Z_API = Contact_API_API.GetGridFunction('Particle Z-Coordinate').GetArray(time_step = Post_TS)
                        Matrix_Contact_API = np.array([X_API, Y_API, Z_API])
                        X_Carrier = Contact_to_1.GetGridFunction('Particle X-Coordinate').GetArray(time_step = Post_TS)
                        Y_Carrier = Contact_to_1.GetGridFunction('Particle Y-Coordinate').GetArray(time_step = Post_TS)
                        Z_Carrier = Contact_to_1.GetGridFunction('Particle Z-Coordinate').GetArray(time_step = Post_TS)
                        Matrix_Contact_Carrier = np.array([X_Carrier, Y_Carrier, Z_Carrier])
                        Matrix_Carrier = Matrix_Contact_Carrier.T
                        Matrix_API = Matrix_Contact_API.T
                        def row_in_matrix(row, matrix):
                            if matrix.size == 0:
                                return False
                            return np.any(np.all(row == matrix, axis=1))
                        Total_API = []
                        for i, row in enumerate(Matrix_Carrier):
                            if not row_in_matrix(row, Matrix_API):
                                Total_API.append(i)
                        PostAPI = len(Total_API)
                        Particle_Mass_Process.SetCutValue(1)
                        API_Array = []
                        for mass in range(0,PostAPI):
                            API_Mass = float(Particle_Mass_Process.GetGridFunction('Particle Mass').GetArray(time_step=Post_TS)[mass]) * 10e11
                            API_Array.append(API_Mass)
                        PostAPI_mass = sum(API_Array)
                        Delta_m = PreAPI_mass - PostAPI_mass
                        Detachment_Rate = Delta_m / PreAPI_mass
                        CRD = Delta_m / (2 * Carrier_Mass) #Carrier relative Detachment rate
                        project.CloseProject(False)
                        PreAPI_Array.append(PreAPI)
                        PostAPI_Array.append(PostAPI)
                        PreAPI_mass_Array.append(PreAPI_mass)
                        PostAPI_mass_Array.append(PostAPI_mass)
                        absolute_Detachment_Array.append(Delta_m)
                        relative_Detachment_Array.append(Detachment_Rate)
                        CRD_Array.append(CRD)
                        Last_TS_Array.append(Post_TS)
                        #Berechnung der Statistik
                        os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Results\Particle-Particle\Results')
                        wb = openpyxl.load_workbook('Results_Particle-Particle.xlsx')
                        Indi_sheet = wb['Individual']
                        Indi_sheet['A1'] = 'Geometry'
                        Indi_sheet['B1'] = 'Loading Replicate'
                        column_letter = openpyxl.utils.get_column_letter(n+2)
                        Indi_sheet[column_letter + '1'] = str(n)
                        Mittelwert_Buchstabe = openpyxl.utils.get_column_letter(n+3)
                        Stab_Buchstabe = openpyxl.utils.get_column_letter(n+4)
                        RSD_Buchstabe = openpyxl.utils.get_column_letter(n+5)
                        Indi_sheet[Mittelwert_Buchstabe + '1'] = 'Mean'
                        Indi_sheet[Stab_Buchstabe + '1'] = 'St.dv.'
                        Indi_sheet[RSD_Buchstabe + '1'] = 'RSD'
                        Indi_sheet['A' + str(Indi_row)] = str(CarrierGeometry)
                        Indi_sheet['B' + str(Indi_row)] = str(Loading_Replicate)
                        Indi_sheet[column_letter + str(Indi_row)].number_format = '0.00%'
                        Indi_sheet[column_letter + str(Indi_row)] = Detachment_Rate
                        #zellen = sheet['C' + str(Indi_row):column_letter + str(Indi_row)]
                        #werte = [zelle.value for reihe in zellen for zelle in reihe if isinstance(zelle.value, (int, float))]
                        Indi_Array.append(Detachment_Rate)
                        mittelwert = np.mean(Indi_Array)
                        stdabw = np.std(Indi_Array, ddof=1)  # StichprobeÊ
                        rel_stdabw = (stdabw / mittelwert) if mittelwert != 0 else None
                        Indi_sheet[Mittelwert_Buchstabe + str(Indi_row)].number_format = '0.00%'
                        Indi_sheet[Mittelwert_Buchstabe + str(Indi_row)] = mittelwert
                        Indi_sheet[Stab_Buchstabe + str(Indi_row)].number_format = '0.00%'
                        Indi_sheet[Stab_Buchstabe + str(Indi_row)] = stdabw
                        Indi_sheet[RSD_Buchstabe + str(Indi_row)].number_format = '0.00%'
                        Indi_sheet[RSD_Buchstabe + str(Indi_row)] = rel_stdabw
                        wb.save('Results_Particle-Particle.xlsx')
                    if Decision == 'no':
                        project.CloseProject(False)
            os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Results\Particle-Particle\Results')
            wb = openpyxl.load_workbook('Results_Particle-Particle.xlsx')
            sheet = wb['Raw']
            rD_RSD = np.std(relative_Detachment_Array, ddof=1) / np.mean(relative_Detachment_Array)
            sheet['A' + str(Excel_row)] = str(Run_row)
            sheet['B' + str(Excel_row)] = np.mean(PreAPI_Array)
            sheet['C' + str(Excel_row)].number_format = '0'
            sheet['C' + str(Excel_row)] = np.mean(PostAPI_Array)
            sheet['D' + str(Excel_row)].number_format = '0.00'
            sheet['D' + str(Excel_row)] = np.mean(PreAPI_mass_Array)
            sheet['E' + str(Excel_row)].number_format = '0.00'
            sheet['E' + str(Excel_row)] = np.mean(PostAPI_mass_Array)
            sheet['F' + str(Excel_row)].number_format = '0.00'
            sheet['F' + str(Excel_row)] = np.mean(absolute_Detachment_Array)
            sheet['G' + str(Excel_row)].number_format = '0.00%'
            sheet['G' + str(Excel_row)] = np.mean(relative_Detachment_Array)
            sheet['H' + str(Excel_row)].number_format = '0.00%'
            sheet['H' + str(Excel_row)] = rD_RSD
            sheet['I' + str(Excel_row)].number_format = '0.00%'
            sheet['I' + str(Excel_row)] = np.mean(CRD_Array)
            sheet['J' + str(Excel_row)] = np.mean(Last_TS_Array)
            wb.save('Results_Particle-Particle.xlsx')
            #DoE eintragen
            os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\DoE')
            while True:
                now = datetime.now()
                if 30 <= now.second <= 40:
                    break
                time.sleep(2)
            wb = openpyxl.load_workbook('DoE_Arrays.xlsx')
            sheet_parameter = wb['Parameter']
            sheet_results = wb['Result']
            for row in sheet_parameter.iter_rows():
                for cell in row:
                    # Schreibe den Wert an dieselbe Position im "Results"-Sheet
                    sheet_results[cell.coordinate].value = cell.value
            for column in range(1,sheet_results.max_column+1):
                letter = openpyxl.utils.get_column_letter(column)
                if sheet_results[letter + str(1)].value == 'PP rD':
                    break
            sheet_results[letter + str(Excel_row)].number_format = '0.00%'
            sheet_results[letter + str(Excel_row)] = np.mean(relative_Detachment_Array)
            wb.save('DoE_Arrays.xlsx')
            #Process eintragen
            os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
            wb = openpyxl.load_workbook('Carrier.xlsx')
            Carriersheet = wb['Carrier']
            cell = Carriersheet['E' + str(Excel_row)]
            cell.font = cell.font.copy(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            Carriersheet['E' + str(Excel_row)] = 'X'
            wb.save('Carrier.xlsx')
    #Spaltenbreite an Text anpassen
    os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\\Particle-Particle\\Results')
    wb = openpyxl.load_workbook('Results_Particle-Particle.xlsx')
    sheet = wb['Raw']
    for spalte in sheet.columns:
        max_laenge = 0
        spaltenbuchstabe = openpyxl.utils.get_column_letter(spalte[0].column)
        max_laenge = len(str(spalte[0].value))
        adjusted_width = (max_laenge + 2) * 1
        if spaltenbuchstabe == 'H':
            adjusted_width *= 1.7
        sheet.column_dimensions[spaltenbuchstabe].width = adjusted_width
    wb.save('Results_Particle-Particle.xlsx')

def ParticleWall_Analyse():
    global Array_dict
    global P_Array_dict
    global Run_Opt
    global Range
    #PostParticle-Wall
    Process = 'ParticleWall'
    for ParaPosition in range(1,P_Array_lenght + 2):
        if P_Array_dict['Parameter'][ParaPosition - 1] == Process:
            Position = ParaPosition - 1
            break
    os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Particle-Wall\\Results')
    if not os.path.exists('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\\Particle-Wall\\Results\\Results_Particle-Wall.xlsx'):
        wb = openpyxl.Workbook()
        wb.save('Results_Particle-Wall.xlsx')
    wb = openpyxl.load_workbook('Results_Particle-Wall.xlsx')
    wb.save('Results_Particle-Wall.xlsx')
    sheet = wb.active
    sheet.title = 'Raw'
    sheet = wb['Raw']
    sheet['A1'] = 'No'
    sheet['B1'] = 'PreAPI'
    sheet['C1'] = 'PostAPI'
    sheet['D1'] = 'PreAPI mass [ng]'
    sheet['E1'] = 'PostAPI mass [ng]'
    sheet['F1'] = 'absolute Detachment [ng]'
    sheet['G1'] = 'relative Detachment [%]'
    sheet['H1'] = 'RSD [%]'
    sheet['I1'] = 'CRD'
    sheet['J1'] = 'Last TS'   
    wb.save('Results_Particle-Wall.xlsx')
    Individual_Name = 'Individual'
    if Individual_Name not in wb.sheetnames:
        wb.create_sheet('Individual')
        Indi_sheet = wb['Individual']
    else:
        Indi_sheet = wb['Individual']
    wb.save('Results_Particle-Wall.xlsx')
    global Simulation_Letter
    Simulation_Letter = 'G'
    Span()
    for Excel_row in Range:
        Array_row = Excel_row - 2
        Run_row = Excel_row - 1
        os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
        wb = openpyxl.load_workbook('Carrier.xlsx')
        Carriersheet = wb['Carrier']
        if Carriersheet['F' + str(Excel_row)].value != None and Carriersheet['G' + str(Excel_row)].value == None:
            CarrierGeometry = str(Carriersheet['A' + str(Excel_row)].value)
            #Define Parameters
            for Parameter in P_Array_dict:
                Value = P_Array_dict[Parameter][Position]
                if Value == 'DoE':
                    DoE_Array_Value = Array_dict[Parameter]
                    DoE_Value = DoE_Array_Value[Array_row]
                    P_Array_dict[Parameter][Position] = DoE_Value
            PreAPI_Array = []
            PostAPI_Array = []
            PreAPI_mass_Array = []
            PostAPI_mass_Array = []
            absolute_Detachment_Array = []
            relative_Detachment_Array = []
            CRD_Array = []
            Last_TS_Array = []
            #Finished Simulation Check
            for Loading_Replicate in range(1,Loading_Replicas+1):
                os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Results\Particle-Wall\Results')
                wb = openpyxl.load_workbook('Results_Particle-Wall.xlsx')
                Indi_sheet = wb['Individual']
                #Indi_Row
                Indi_Array = []
                for row in range(2,Indi_sheet.max_row+2):
                    if Indi_sheet['A' + str(row)].value == None:
                        Indi_row = row
                        break
                for n in range(1,n_max):
                    filepath_PPSimulations = 'D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Particle-Wall\\Simulations'
                    os.chdir(filepath_PPSimulations)
                    filename_PPSimulation = str(Run_row) + '_' + CarrierGeometry + '_PW-Collision' + '_L' + str(Loading_Replicate) +  '_C' + str(n) + '.rocky'
                    app.OpenProject(filename_PPSimulation)   
                    study = api.GetStudy()
                    Finished = study.HasResults()
                    project = api.GetProject()
                    if Finished == False:
                        Decision = 'no'
                    elif Finished == True:
                        Decision = 'yes'
                    #Bei Entscheidung Ja
                    if Decision == 'yes':
                        #Projekt
                        Pre_TS = 1
                        project = app.GetProject()
                        study = api.GetStudy()
                        user_processes = project.GetUserProcessCollection()
                        particles = study.GetParticles()
                        Contacts = study.GetContactData()
                        #Carrier
                        CarrierPosition = user_processes.CreatePropertyProcess(particles, 'CarrierPosition')
                        CarrierPosition.SetPropertyGridFunction('Particle Group')
                        CarrierPosition.SetCutValue(0)
                        #Create Cube
                        Get_Carrier_Process = user_processes.GetProcess('CarrierPosition')
                        Cube_Carrier = user_processes.CreateCubeProcess(Get_Carrier_Process, 'Cube_Carrier')
                        Cube_Carrier.SetCenter(0.0, 0.0, 0.0, 'm')
                        Cube_Carrier.SetSize(20.0, 10.0, 10.0, 'mm')
                        #Create Eulerian
                        Eulerian_Carrier = user_processes.CreateEulerianStatistics(Cube_Carrier, 'Eulerian_Carrier')
                        Eulerian_Carrier.SetDivisions((1, 1, 1,))
                        #Contact from Carrier
                        From_Carrier = user_processes.CreateContactToParticleProcess(Contacts, 'To Carrier')
                        #API
                        #Contact to Particle
                        Contact_to_1 = user_processes.CreatePropertyProcess(From_Carrier, 'Contact_to_1')
                        Contact_to_1.SetPropertyGridFunction('Particle Group')
                        Contact_to_1.SetCutValue(1)
                        #Create Cube
                        Get_Contact_to_1 = user_processes.GetProcess('Contact_to_1')
                        Cube_1 = user_processes.CreateCubeProcess(Get_Contact_to_1, 'Cube_1')
                        Cube_1.SetCenter(0.0, 0.0, 0.0, 'm')
                        Cube_1.SetSize(20.0, 10.0, 10.0, 'mm')
                        #Create Eulerian
                        Eulerian_1 = user_processes.CreateEulerianStatistics(Cube_1, 'Eulerian_1')
                        Eulerian_1.SetDivisions((1, 1, 1,))
                        #API to API
                        API_to_API = user_processes.CreateParticleToContactProcess(Contact_to_1, 'From_API')
                        From_API_to_API = user_processes.CreatePropertyProcess(API_to_API, 'From_API_to_API')
                        From_API_to_API.SetPropertyGridFunction('Particle Group From')
                        From_API_to_API.SetCutValue(1)
                        contact_API_API = user_processes.CreateContactToParticleProcess(From_API_to_API, 'contact_API_API')
                        Contact_API_API = user_processes.CreatePropertyProcess(contact_API_API, 'Contact_API_API')
                        Contact_API_API.SetPropertyGridFunction('Particle Group')
                        Contact_API_API.SetCutValue(1)
                        #Create Cube
                        Get_Contact_to_API = user_processes.GetProcess('Contact_API_API')
                        Cube_API = user_processes.CreateCubeProcess(Get_Contact_to_API, 'Cube_API')
                        Cube_API.SetCenter(0.0, 0.0, 0.0, 'm')
                        Cube_API.SetSize(20.0, 10.0, 10.0, 'mm')
                        #Create Eulerian
                        Eulerian_API = user_processes.CreateEulerianStatistics(Cube_API, 'Eulerian_API')
                        Eulerian_API.SetDivisions((1, 1, 1,))
                        #Write Results
                        os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Particle-Wall\\Results')
                        wb = openpyxl.load_workbook('Results_Particle-Wall.xlsx')
                        sheet = wb['Raw']
                        Particle_Mass_Process = user_processes.CreatePropertyProcess(particles)
                        Particle_Mass_Process.SetName('ParticleMassProcess')
                        Particle_Mass_Process.SetPropertyGridFunction('Particle Group')
                        Particle_Mass_Process.SetCutValue(0)
                        Carrier_Mass = float(Particle_Mass_Process.GetGridFunction('Particle Mass').GetArray(time_step=1)[0]) * 10e11
                        Particle_Mass_Process.SetCutValue(1)
                        API_Mass = float(Particle_Mass_Process.GetGridFunction('Particle Mass').GetArray(time_step=1)[0]) * 10e11
                        #API in Kontakt und 2. Schicht zusammenführen
                        X_API = Contact_API_API.GetGridFunction('Particle X-Coordinate').GetArray(time_step = Pre_TS)
                        Y_API = Contact_API_API.GetGridFunction('Particle Y-Coordinate').GetArray(time_step = Pre_TS)
                        Z_API = Contact_API_API.GetGridFunction('Particle Z-Coordinate').GetArray(time_step = Pre_TS)
                        Matrix_Contact_API = np.array([X_API, Y_API, Z_API])
                        X_Carrier = Contact_to_1.GetGridFunction('Particle X-Coordinate').GetArray(time_step = Pre_TS)
                        Y_Carrier = Contact_to_1.GetGridFunction('Particle Y-Coordinate').GetArray(time_step = Pre_TS)
                        Z_Carrier = Contact_to_1.GetGridFunction('Particle Z-Coordinate').GetArray(time_step = Pre_TS)
                        Matrix_Contact_Carrier = np.array([X_Carrier, Y_Carrier, Z_Carrier])
                        Matrix_Carrier = Matrix_Contact_Carrier.T
                        Matrix_API = Matrix_Contact_API.T
                        def row_in_matrix(row, matrix):
                            if matrix.size == 0:
                                return False
                            return np.any(np.all(row == matrix, axis=1))
                        Total_API = []
                        if not Matrix_API.size == 0:
                            for i, row in enumerate(Matrix_API):
                                if not row_in_matrix(row, Matrix_Carrier):
                                    Total_API.append(i)
                        PreAPI = len(X_Carrier) + len(Total_API)
                        #PreAPI = len(Total_API)
                        Particle_Mass_Process.SetCutValue(1)
                        API_Array = []
                        for mass in range(0,PreAPI):
                            API_Mass = float(Particle_Mass_Process.GetGridFunction('Particle Mass').GetArray(time_step=Pre_TS)[mass]) * 10e11
                            API_Array.append(API_Mass)
                        PreAPI_mass = sum(API_Array)
                        SimulatorRun = study.GetSimulatorRun()
                        OutFre = SimulatorRun.GetOutputFrequency()
                        SimDur = SimulatorRun.GetSimulationDuration()
                        Last_TS = int(SimDur / OutFre)
                        a = float(Eulerian_1.GetGridFunction('Number of Particles').GetArray(time_step = 3))
                        Post_TS = Last_TS
                        for TS in range(4,Last_TS):
                            b = float(Eulerian_1.GetGridFunction('Number of Particles').GetArray(time_step = TS))
                            if b < a:
                                Post_TS = TS
                                break
                            a = b
                        if Post_TS != Last_TS:
                            a = float(Eulerian_1.GetGridFunction('Number of Particles').GetArray(time_step = Post_TS + 1))
                            for TS in range(Post_TS + 2, Last_TS):
                                b = float(Eulerian_1.GetGridFunction('Number of Particles').GetArray(time_step = TS))
                                if b >= a:
                                    Post_TS = TS + 5
                                    break
                                a = b
                        if Post_TS > Last_TS:
                            Post_TS = Last_TS
                        Carrier_Elimination = 100
                        for timepoint in range(1,Last_TS):
                            if int(Eulerian_Carrier.GetGridFunction('Number of Particles').GetArray(time_step = timepoint)) == 0:
                                Carrier_Elimination = timepoint
                                break
                        if Post_TS > Carrier_Elimination:
                            Post_TS = Carrier_Elimination - 5
                        #Individual Post_TS
                        Post_TS = 40
                        #API in Kontakt und 2. Schicht zusammenführen
                        X_API = Contact_API_API.GetGridFunction('Particle X-Coordinate').GetArray(time_step = Post_TS)
                        Y_API = Contact_API_API.GetGridFunction('Particle Y-Coordinate').GetArray(time_step = Post_TS)
                        Z_API = Contact_API_API.GetGridFunction('Particle Z-Coordinate').GetArray(time_step = Post_TS)
                        Matrix_Contact_API = np.array([X_API, Y_API, Z_API])
                        X_Carrier = Contact_to_1.GetGridFunction('Particle X-Coordinate').GetArray(time_step = Post_TS)
                        Y_Carrier = Contact_to_1.GetGridFunction('Particle Y-Coordinate').GetArray(time_step = Post_TS)
                        Z_Carrier = Contact_to_1.GetGridFunction('Particle Z-Coordinate').GetArray(time_step = Post_TS)
                        Matrix_Contact_Carrier = np.array([X_Carrier, Y_Carrier, Z_Carrier])
                        Matrix_Carrier = Matrix_Contact_Carrier.T
                        Matrix_API = Matrix_Contact_API.T
                        def row_in_matrix(row, matrix):
                            if matrix.size == 0:
                                return False
                            return np.any(np.all(row == matrix, axis=1))
                        Total_API = []
                        for i, row in enumerate(Matrix_Carrier):
                            if not row_in_matrix(row, Matrix_API):
                                Total_API.append(i)
                        PostAPI = len(Total_API)
                        Particle_Mass_Process.SetCutValue(1)
                        API_Array = []
                        for mass in range(0,PostAPI):
                            API_Mass = float(Particle_Mass_Process.GetGridFunction('Particle Mass').GetArray(time_step=Post_TS)[mass]) * 10e11
                            API_Array.append(API_Mass)
                        PostAPI_mass = sum(API_Array)
                        Delta_m = PreAPI_mass - PostAPI_mass
                        Detachment_Rate = Delta_m / PreAPI_mass
                        CRD = Delta_m / Carrier_Mass #Carrier relative Detachment rate
                        project.CloseProject(False)
                        PreAPI_Array.append(PreAPI)
                        PostAPI_Array.append(PostAPI)
                        PreAPI_mass_Array.append(PreAPI_mass)
                        PostAPI_mass_Array.append(PostAPI_mass)
                        absolute_Detachment_Array.append(Delta_m)
                        relative_Detachment_Array.append(Detachment_Rate)
                        CRD_Array.append(CRD)
                        Last_TS_Array.append(Post_TS)
                        #Berechnung der Statistik
                        os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Results\Particle-Wall\Results')
                        wb = openpyxl.load_workbook('Results_Particle-Wall.xlsx')
                        Indi_sheet = wb['Individual']
                        Indi_sheet['A1'] = 'Geometry'
                        Indi_sheet['B1'] = 'Loading Replicate'
                        column_letter = openpyxl.utils.get_column_letter(n+2)
                        Indi_sheet[column_letter + '1'] = str(n)
                        Mittelwert_Buchstabe = openpyxl.utils.get_column_letter(n+3)
                        Stab_Buchstabe = openpyxl.utils.get_column_letter(n+4)
                        RSD_Buchstabe = openpyxl.utils.get_column_letter(n+5)
                        Indi_sheet[Mittelwert_Buchstabe + '1'] = 'Mean'
                        Indi_sheet[Stab_Buchstabe + '1'] = 'St.dv.'
                        Indi_sheet[RSD_Buchstabe + '1'] = 'RSD'
                        Indi_sheet['A' + str(Indi_row)] = str(CarrierGeometry)
                        Indi_sheet['B' + str(Indi_row)] = str(Loading_Replicate)
                        Indi_sheet[column_letter + str(Indi_row)].number_format = '0.00%'
                        Indi_sheet[column_letter + str(Indi_row)] = Detachment_Rate
                        Indi_Array.append(Detachment_Rate)
                        mittelwert = np.mean(Indi_Array)
                        stdabw = np.std(Indi_Array, ddof=1)  # StichprobeÊ
                        rel_stdabw = (stdabw / mittelwert) if mittelwert != 0 else None
                        Indi_sheet[Mittelwert_Buchstabe + str(Indi_row)].number_format = '0.00%'
                        Indi_sheet[Mittelwert_Buchstabe + str(Indi_row)] = mittelwert
                        Indi_sheet[Stab_Buchstabe + str(Indi_row)].number_format = '0.00%'
                        Indi_sheet[Stab_Buchstabe + str(Indi_row)] = stdabw
                        Indi_sheet[RSD_Buchstabe + str(Indi_row)].number_format = '0.00%'
                        Indi_sheet[RSD_Buchstabe + str(Indi_row)] = rel_stdabw
                        wb.save('Results_Particle-Wall.xlsx')               
                if Decision == 'no':
                    project.CloseProject(False)
                os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Results\Particle-Wall\Results')
                wb = openpyxl.load_workbook('Results_Particle-Wall.xlsx')
                sheet = wb['Raw']
                rD_RSD = np.std(relative_Detachment_Array, ddof=1) / np.mean(relative_Detachment_Array)
                sheet['A' + str(Excel_row)] = str(Run_row)
                sheet['B' + str(Excel_row)] = np.mean(PreAPI_Array)
                sheet['C' + str(Excel_row)].number_format = '0'
                sheet['C' + str(Excel_row)] = np.mean(PostAPI_Array)
                sheet['D' + str(Excel_row)].number_format = '0.00'
                sheet['D' + str(Excel_row)] = np.mean(PreAPI_mass_Array)
                sheet['E' + str(Excel_row)].number_format = '0.00'
                sheet['E' + str(Excel_row)] = np.mean(PostAPI_mass_Array)
                sheet['F' + str(Excel_row)].number_format = '0.00'
                sheet['F' + str(Excel_row)] = np.mean(absolute_Detachment_Array)
                sheet['G' + str(Excel_row)].number_format = '0.00%'
                sheet['G' + str(Excel_row)] = np.mean(relative_Detachment_Array)
                sheet['H' + str(Excel_row)].number_format = '0.00%'
                sheet['H' + str(Excel_row)] = rD_RSD
                sheet['I' + str(Excel_row)].number_format = '0.00%'
                sheet['I' + str(Excel_row)] = np.mean(CRD_Array)
                sheet['J' + str(Excel_row)] = np.mean(Last_TS_Array)
                wb.save('Results_Particle-Wall.xlsx')
                #DoE eintragen
                os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\DoE')
                while True:
                    now = datetime.now()
                    if 30 <= now.second <= 40:
                        break
                    time.sleep(2)
                wb = openpyxl.load_workbook('DoE_Arrays.xlsx')
                sheet_parameter = wb['Parameter']
                sheet_results = wb['Result']
                for row in sheet_parameter.iter_rows():
                    for cell in row:
                        # Schreibe den Wert an dieselbe Position im "Results"-Sheet
                        sheet_results[cell.coordinate].value = cell.value
                for column in range(1,sheet_results.max_column+1):
                    letter = openpyxl.utils.get_column_letter(column)
                    if sheet_results[letter + str(1)].value == 'PW rD':
                        break
                sheet_results[letter + str(Excel_row)].number_format = '0.00%'
                PP_letter = openpyxl.utils.get_column_letter(column-1)
                Mean_letter = openpyxl.utils.get_column_letter(column+1)
                PW_rD = np.mean(relative_Detachment_Array)
                sheet_results[letter + str(Excel_row)] = PW_rD
                PP_rD = sheet_results[PP_letter + str(Excel_row)].value
                Mean_rD = (PW_rD + PP_rD) / 2
                sheet_results[Mean_letter + str(Excel_row)].number_format = '0.00%'
                sheet_results[Mean_letter + str(Excel_row)] = Mean_rD
                wb.save('DoE_Arrays.xlsx')
                #Process eintragen
                os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
                wb = openpyxl.load_workbook('Carrier.xlsx')
                Carriersheet = wb['Carrier']
                cell = Carriersheet['G' + str(Excel_row)]
                cell.font = cell.font.copy(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                Carriersheet['G' + str(Excel_row)] = 'X'
                wb.save('Carrier.xlsx')
    #Spaltenbreite an Text anpassen
    os.chdir('D:\Rocky_Simulations\\Rocky AP-'  + Path_Number + '\\Results\\Particle-Wall\\Results')
    wb = openpyxl.load_workbook('Results_Particle-Wall.xlsx')
    sheet = wb['Raw']
    for spalte in sheet.columns:
        max_laenge = 0
        spaltenbuchstabe = openpyxl.utils.get_column_letter(spalte[0].column)
        max_laenge = len(str(spalte[0].value))
        adjusted_width = (max_laenge + 2) * 1
        if spaltenbuchstabe == 'H':
            adjusted_width *= 1.7
        sheet.column_dimensions[spaltenbuchstabe].width = adjusted_width
    wb.save('Results_Particle-Wall.xlsx')

def BO_Loop():
    global FollowUp
    global Current_Iterations
    Current_Iterations = 0 ###Hot Fix?
    while Current_Iterations <= max_iterations:
        if FollowUp == False:
            break
        if FollowUp == True:
            Update_Iterations()
            DoE_Read()
            Parameter_Read()
            ###Span für alle Simulationstypen vergleichen
            os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')
            wb = openpyxl.load_workbook('Carrier.xlsx')
            Carriersheet = wb['Carrier']
            Range_Array = []
            for col in Carriersheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=Carriersheet.max_column):
                col_index = col[0].column  # erste Zelle der Spalte → Spaltennummer
                global Simulation_Letter
                Simulation_Letter = openpyxl.utils.get_column_letter(col_index)
                #Simulation_Letter = 'A'
                Span()
                Range_Array.append(np.size(Range))
            Max_Range = max(Range_Array)
            if Max_Range > 0:#np.size(Range) > 0:
                ##Durchführen
                Loading_Create()
                Loading_Analyse()
                ParticleParticle_Create()
                ParticleWall_Create()
                ParticleParticle_Analyse()
                ParticleWall_Analyse()
            else:
                ###Time Pending
                time.sleep(30)
                #input('Break')
                #break
                Break_Check()
                global Break_True
                if Break_True == True:
                    break
                    os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\Process')     
                    wb = openpyxl.load_workbook('Recent.xlsx')
                    sheet = wb['Processes']
                    sheet['B10'] = 0
                    wb.save('Recent.xlsx')
                if not FollowUp: 
                    break


BO_Loop()

input('Ende')

    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    