# -*- coding: utf-8 -*-
"""
Created on Tue May 20 06:52:04 2025

@author: sunpa675
"""

import os
import openpyxl
import numpy as np
import GPyOpt
import time
from datetime import datetime

def Path_Iteration_Check():
    os.chdir('D:\Rocky_Simulations\Empty_Simulation_Folder\Process')
    wb = openpyxl.load_workbook('Recent.xlsx')
    sheet = wb['Path']
    global max_iterations
    max_iterations = sheet['A2'].value
    global Path_Number
    Path_Number_long = sheet['A1'].value
    number_str = str(Path_Number_long)
    Path_Number = ".".join(number_str)

Path_Iteration_Check()
os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\DoE')

def read_Range_domain(wb, sheet_name='Range'):
    os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\DoE')
    wb = openpyxl.load_workbook('DoE_Arrays.xlsx')
    global domain
    sheet = wb[sheet_name]
    domain = []
    row = 2
    while True:
        param_name = sheet[f'A{row}'].value
        if param_name is None:
            break
        param_min = sheet[f'B{row}'].value
        param_max = sheet[f'C{row}'].value
        if param_min is None or param_max is None:
            print(f"Warnung: Min oder Max für Parameter {param_name} fehlen in Zeile {row}")
            row += 1
            continue
        domain.append({
            'name': param_name,
            'type': 'continuous',
            'domain': (float(param_min), float(param_max))
        })
        row += 1
    return domain


def read_data(wb):
    Resultsheet = wb['Result']
    Parametersheet = wb['Parameter']
    #Rangesheet = wb['Range']
    global domain
    #global x_werte_array
    #x_werte_array = []
    start_col = 2
    end_col = start_col + len(domain)-1
    #global data
    data = []
    for row in Resultsheet.iter_rows(min_row=2, min_col=start_col, max_col=end_col):
        row_values = [cell.value for cell in row]
        data.append(row_values)
    # Umwandeln in NumPy-Array
    global matrix
    matrix = np.array(data, dtype=float)    
    # Zwei Y-Werte aus Spalten E, F (also 5, 6)
    y1_werte = [row[0].value for row in Resultsheet.iter_rows(min_row=2, min_col=len(domain)+2, max_col=len(domain)+2)]
    y2_werte = [row[0].value for row in Resultsheet.iter_rows(min_row=2, min_col=len(domain)+3, max_col=len(domain)+3)]
    # Länge von x array aus Parametersheet
    xP_werte = [row[0].value for row in Parametersheet.iter_rows(min_row=2, min_col=2, max_col=2)]
    # Filtere nur vollständige Daten (keine None)
    valid_indices = [i for i in range(len(matrix)) 
                 if None not in matrix[i, :] 
                 and None not in (y1_werte[i], y2_werte[i])]
    if not valid_indices:
        return None, None, Parametersheet, xP_werte
    #X_real = np.array([[x1_werte[i], x2_werte[i], x3_werte[i], x4_werte[i]] for i in valid_indices])
    X_real = matrix
    Y_raw = np.array([[y1_werte[i], y2_werte[i]] for i in valid_indices])
    # Mittelwert der beiden Y-Werte
    Y_real = Y_raw.mean(axis=1).reshape(-1,1)
    return X_real, Y_real, Parametersheet, xP_werte

def write_next_parameter(wb, Parametersheet, x_next):
    os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\DoE')
    global domain
    wb = openpyxl.load_workbook('DoE_Arrays.xlsx')
    Parametersheet = wb['Parameter']
    for row in range(2, Parametersheet.max_row + 10):
        if Parametersheet[f'B{row}'].value is None:
            Parametersheet[f'A{row}'] = int(row-1)
            for true_row in range(2,len(domain)+2):
                column_letter = openpyxl.utils.get_column_letter(true_row)
                Parametersheet[column_letter + str(row)] = float(x_next[0, true_row-2])
            wb.save('DoE_Arrays.xlsx')
            print(f'Neuer Parameter geschrieben in Zeile {row}')
            return True
    print('Keine freie Zeile gefunden zum Schreiben!')
    return False

def get_best_result(X_real, Y_real):
    best_index = np.argmax(Y_real)
    best_x = X_real[best_index, :]  # alle 3 x-Werte als Array
    best_y = float(Y_real[best_index, 0])
    return best_x, best_y

last_result_count = 0
print("Starte Polling...")

while True:
    X_real = []
    Y_real = []
    #Auf bestimmte Sekunden-Zeit einstellen
    while True:
        now = datetime.now()
        if 0 <= now.second <= 10:
            break
        time.sleep(2)
    wb = openpyxl.load_workbook('DoE_Arrays.xlsx')
    read_Range_domain(wb, 'Range')
    Path_Iteration_Check()
    os.chdir('D:\Rocky_Simulations\Rocky AP-'  + Path_Number + '\DoE')
    wb = openpyxl.load_workbook('DoE_Arrays.xlsx')
    data = read_data(wb)    
    # Fall 1: read_data gibt None zurück → weiter warten
    if data is None:
        print("Noch keine gültigen Daten vorhanden, warte...")
        time.sleep(50)
        continue
    # Fall 2: read_data gibt etwas zurück, aber X oder Y fehlen
    X_real, Y_real, Parametersheet, xP_werte = data
    if X_real is None or Y_real is None:
        print("Noch keine vollständigen X- und Y-Daten vorhanden, warte...")
        time.sleep(50)
        continue
    # Ab hier ist sicher: X_real und Y_real sind numpy-Arrays
    Start_Value = 3
    if len(Y_real) >= Start_Value and len(Y_real) == len(xP_werte):# and len(Y_real) > last_result_count:
        print(f"Neue Ergebnisse erkannt: {len(Y_real)} Messungen insgesamt.")
        wb = openpyxl.load_workbook('DoE_Arrays.xlsx')
        domain = read_Range_domain(wb, 'Range')
        bo = GPyOpt.methods.BayesianOptimization(f=None, domain=domain, X=X_real, Y=-Y_real)
        x_next = bo.suggest_next_locations()
        write_next_parameter(wb, Parametersheet, x_next)
        best_x, best_y = get_best_result(X_real, Y_real)
        best_y = best_y * 100
        print(f"Bestes Ergebnis bisher: X = {best_x} mit Detachment Rate = {best_y:.3f}")
        last_result_count = len(Y_real)
    elif len(Y_real) >= max_iterations:
        best_x, best_y = get_best_result(X_real, Y_real)
        best_y = best_y * 100
        print(f"\nMaximale Iterationen ({max_iterations}) erreicht.")
        break
    else:
        print("Keine neuen Ergebnisse erkannt, warte...")
    time.sleep(50)




